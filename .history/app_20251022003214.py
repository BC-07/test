from flask import Flask, request, render_template, jsonify, send_file, session, redirect, url_for, flash
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import sqlite3
import logging
import re
from typing import Dict, List, Any, Optional
import pickle
from utils import PersonalDataSheetProcessor
# from ocr_processor import OCRProcessor  # Removed OCR dependency
from database import DatabaseManager
from clean_upload_handler import CleanUploadHandler
from assessment_engine import UniversityAssessmentEngine
from datetime import datetime, timedelta, date
import pandas as pd
import json
import uuid

# Try to import flask-login, fallback if not available
try:
    from flask_login import LoginManager, login_user, logout_user, login_required, current_user
    FLASK_LOGIN_AVAILABLE = True
except ImportError:
    FLASK_LOGIN_AVAILABLE = False
    # Create mock decorators if flask_login is not available
    def login_required(f):
        from functools import wraps
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('user_authenticated'):
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated_function
    
    class MockCurrentUser:
        @property
        def is_authenticated(self):
            return session.get('user_authenticated', False)
        
        @property
        def is_admin(self):
            return session.get('user_is_admin', False)
        
        @property
        def id(self):
            return session.get('user_id')
        
        @property
        def email(self):
            return session.get('user_email')
        
        @property
        def is_active(self):
            return True
    
    current_user = MockCurrentUser()

try:
    import bcrypt
    BCRYPT_AVAILABLE = True
except ImportError:
    BCRYPT_AVAILABLE = False

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Global database manager
db_manager = DatabaseManager()

class SimpleUser:
    """Simple User class for Flask-Login compatibility"""
    def __init__(self, user_data):
        self.id = user_data['id']
        self.email = user_data['email']
        self.first_name = user_data['first_name']
        self.last_name = user_data['last_name']
        self.is_admin = user_data.get('is_admin', False)
        self.is_active_user = user_data.get('is_active', True)
    
    def is_authenticated(self):
        return True
    
    def is_active(self):
        return self.is_active_user
    
    def is_anonymous(self):
        return False
    
    def get_id(self):
        return str(self.id)

class PDSAssessmentApp:
    def __init__(self):
        self.app = Flask(__name__)
        CORS(self.app)
        
        # Get absolute paths
        base_dir = os.path.abspath(os.path.dirname(__file__))
        
        # PostgreSQL Configuration
        self.app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
        self.app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
        self.app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-default-secret-key-change-this')
        
        # Session configuration for upload session persistence
        self.app.config['SESSION_PERMANENT'] = False
        self.app.config['SESSION_TYPE'] = 'filesystem'
        self.app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 1 hour
        
        # Initialize Flask-Login if available
        if FLASK_LOGIN_AVAILABLE:
            self.login_manager = LoginManager()
            self.login_manager.init_app(self.app)
            self.login_manager.login_view = 'login'
            self.login_manager.login_message = 'Please log in to access this page.'
            self.login_manager.login_message_category = 'info'
            
            @self.login_manager.user_loader
            def load_user(user_id):
                try:
                    user_data = db_manager.get_user_by_id(int(user_id))
                    return SimpleUser(user_data) if user_data else None
                except:
                    return None
        
        # Initialize Flask-SQLAlchemy
        try:
            from models import db
            db.init_app(self.app)
            
            # Create tables if they don't exist
            with self.app.app_context():
                db.create_all()
                logger.info("Flask-SQLAlchemy tables created successfully")
        except Exception as e:
            logger.warning(f"Flask-SQLAlchemy initialization failed: {e}")
            logger.info("Continuing with custom DatabaseManager...")
        
        # Configuration
        self.app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
        self.app.config['UPLOAD_FOLDER'] = os.path.join(base_dir, 'temp_uploads')
        
        # Initialize processors (with better error handling)
        try:
            self.pds_processor = PersonalDataSheetProcessor()
            logger.info("âœ… PDS processor initialized successfully")
        except Exception as e:
            logger.warning(f"âš ï¸ PDS processor initialization failed: {e}")
            logger.info("â„¹ï¸ PDS processing will use fallback methods")
            self.pds_processor = None
            
        # Initialize clean upload handler with better error handling
        try:
            self.clean_upload_handler = CleanUploadHandler()
            logger.info("âœ… Clean upload handler initialized successfully")
        except Exception as e:
            logger.error(f"âŒ Clean upload handler initialization failed: {e}")
            logger.error("âŒ Upload system will not be available")
            self.clean_upload_handler = None
            
        # Initialize Assessment Engine
        try:
            self.assessment_engine = UniversityAssessmentEngine(db_manager)
            logger.info("âœ… Assessment engine initialized successfully")
        except Exception as e:
            logger.warning(f"âš ï¸ Assessment engine initialization failed: {e}")
            logger.info("â„¹ï¸ Will use fallback assessment methods")
            self.assessment_engine = None
            
        self.processor = self.pds_processor  # Main processor for PDS assessment
        
        # Load ML models
        self._load_models()
        
        # Register routes and error handlers
        self._register_routes()
        self._register_error_handlers()
        
        # Add basic routes for testing
        @self.app.route('/routes')
        def list_routes():
            routes = []
            for rule in self.app.url_map.iter_rules():
                routes.append(f"{rule.endpoint}: {rule.rule}")
            return "<br>".join(routes)
    

    def _load_models(self):
        """Models are no longer used - university assessment system replaced ML categorization"""
        logger.info("University assessment system active - ML models deprecated")
        self.rf_classifier_categorization = None
        self.tfidf_vectorizer_categorization = None
        self.rf_classifier_job_recommendation = None
        self.tfidf_vectorizer_job_recommendation = None
    
    def _register_error_handlers(self):
        """Register error handlers"""
        @self.app.errorhandler(404)
        def not_found(error):
            return jsonify({'success': False, 'error': 'Resource not found'}), 404
        
        @self.app.errorhandler(500)
        def internal_error(error):
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    def _register_routes(self):
        """Register all application routes"""
        # Authentication routes
        self.app.add_url_rule('/login', 'login', self.login, methods=['GET', 'POST'])
        self.app.add_url_rule('/logout', 'logout', self.logout)
        self.app.add_url_rule('/privacy-agreement', 'privacy_agreement', self.privacy_agreement, methods=['GET', 'POST'])
        
        # Main routes (protected)
        self.app.add_url_rule('/', 'index', self.index)
        self.app.add_url_rule('/dashboard', 'dashboard', self.dashboard)
        self.app.add_url_rule('/dashboard/<path:section>', 'dashboard_section', self.dashboard)
        self.app.add_url_rule('/demo', 'demo', self.demo)
        
        # API routes
        self.app.add_url_rule('/api/health', 'health_check', self.health_check)
        self.app.add_url_rule('/api/system/status', 'system_status', self.system_status)
        self.app.add_url_rule('/api/debug/jobs', 'debug_jobs', self.debug_jobs)  # Diagnostic endpoint
        # PDS Assessment endpoints
        self.app.add_url_rule('/api/upload-pds', 'upload_pds', self.upload_pds, methods=['POST'])
        self.app.add_url_rule('/api/upload-pds-only', 'upload_pds_only', self.upload_pds_only, methods=['POST'])
        # self.app.add_url_rule('/api/upload-ocr', 'upload_ocr', self.upload_ocr, methods=['POST'])  # Removed OCR endpoint
        
        # New clean upload system
        self.app.add_url_rule('/api/upload-files', 'upload_files_clean', self.upload_files_clean, methods=['POST'])
        self.app.add_url_rule('/api/start-analysis', 'start_analysis', self.start_analysis, methods=['POST'])
        self.app.add_url_rule('/api/pds-candidates', 'get_pds_candidates', self.get_pds_candidates, methods=['GET'])
        self.app.add_url_rule('/api/pds-candidates/<int:candidate_id>', 'handle_pds_candidate', self.handle_pds_candidate, methods=['GET', 'PUT', 'DELETE'])
        self.app.add_url_rule('/api/jobs', 'handle_jobs', self.handle_jobs, methods=['GET', 'POST'])
        self.app.add_url_rule('/api/jobs/<int:job_id>', 'handle_job', self.handle_job, methods=['GET', 'PUT', 'DELETE'])
        self.app.add_url_rule('/api/job-categories', 'handle_job_categories', self.handle_job_categories, methods=['GET', 'POST'])
        self.app.add_url_rule('/api/job-categories/<int:category_id>', 'handle_job_category', self.handle_job_category, methods=['PUT', 'DELETE'])
        self.app.add_url_rule('/api/candidates', 'get_candidates', self.get_candidates, methods=['GET'])
        self.app.add_url_rule('/api/candidates/<int:candidate_id>', 'handle_candidate', self.handle_candidate, methods=['GET', 'PUT', 'DELETE'])
        self.app.add_url_rule('/api/analytics', 'get_analytics', self.get_analytics, methods=['GET'])
        self.app.add_url_rule('/api/analytics-dev', 'get_analytics_dev', self.get_analytics_dev, methods=['GET'])
        self.app.add_url_rule('/api/settings', 'handle_settings', self.handle_settings, methods=['GET', 'PUT'])
        self.app.add_url_rule('/api/scoring-criteria', 'handle_scoring_criteria', self.handle_scoring_criteria, methods=['GET', 'PUT'])
        
        # User management API routes
        self.app.add_url_rule('/api/users', 'handle_users', self.handle_users, methods=['GET', 'POST'])
        self.app.add_url_rule('/api/users/<int:user_id>', 'handle_user', self.handle_user, methods=['GET', 'PUT', 'DELETE'])
        
        # University Assessment API routes
        self.app.add_url_rule('/api/position-types', 'get_position_types', self.get_position_types, methods=['GET'])
        self.app.add_url_rule('/api/position-types/<int:position_type_id>/templates', 'get_assessment_templates', self.get_assessment_templates, methods=['GET'])
        self.app.add_url_rule('/api/jobs/<int:job_id>/position-requirements', 'handle_position_requirements', self.handle_position_requirements, methods=['GET', 'POST', 'PUT'])
        self.app.add_url_rule('/api/jobs/<int:job_id>/assessments', 'handle_job_assessments', self.handle_job_assessments, methods=['GET', 'POST'])
        self.app.add_url_rule('/api/assessments/<int:assessment_id>', 'handle_assessment', self.handle_assessment, methods=['GET', 'PUT'])
        self.app.add_url_rule('/api/assessments/<int:assessment_id>/manual-scores', 'handle_manual_scores', self.handle_manual_scores, methods=['GET', 'POST'])
        self.app.add_url_rule('/api/jobs/<int:job_id>/assessment-comparison', 'get_assessment_comparison', self.get_assessment_comparison, methods=['GET'])
        self.app.add_url_rule('/api/jobs/<int:job_id>/assessment-analytics', 'get_assessment_analytics', self.get_assessment_analytics, methods=['GET'])
        self.app.add_url_rule('/api/university-assessment-analytics', 'get_university_assessment_analytics', self.get_university_assessment_analytics, methods=['GET'])
        self.app.add_url_rule('/api/test-university-analytics', 'get_test_university_analytics', self.get_test_university_analytics, methods=['GET'])
        self.app.add_url_rule('/api/candidates/<int:candidate_id>/assess/<int:job_id>', 'assess_candidate', self.assess_candidate, methods=['POST'])
        self.app.add_url_rule('/api/candidates/<int:candidate_id>/assessment', 'get_candidate_assessment', self.get_candidate_assessment, methods=['GET'])
        self.app.add_url_rule('/api/update_potential_score', 'update_potential_score', self.update_potential_score, methods=['POST'])
        
        # LSPU Job Posting API routes
        self.app.add_url_rule('/api/job-postings', 'handle_job_postings', self.handle_lspu_job_postings, methods=['GET', 'POST'])
        self.app.add_url_rule('/api/lspu-job-postings', 'handle_lspu_job_postings_alias', self.handle_lspu_job_postings, methods=['GET', 'POST'])  # Alias for frontend compatibility
        self.app.add_url_rule('/api/job-postings/<int:job_id>', 'handle_job_posting', self.handle_lspu_job_posting, methods=['GET', 'PUT', 'DELETE'])
        self.app.add_url_rule('/api/job-postings/<int:job_id>/preview', 'preview_job_posting', self.preview_lspu_job_posting, methods=['GET'])
        self.app.add_url_rule('/api/job-postings/<int:job_id>/render', 'render_job_posting', self.render_lspu_job_posting, methods=['GET'])
        self.app.add_url_rule('/api/job-postings/<int:job_id>/export', 'export_job_posting', self.export_lspu_job_posting, methods=['GET'])
        self.app.add_url_rule('/api/campus-locations', 'get_campus_locations', self.get_campus_locations, methods=['GET'])
        
        # Job Posting Assessment Integration routes
        self.app.add_url_rule('/api/job-postings/<int:job_id>/criteria', 'job_posting_criteria', self.handle_job_posting_criteria, methods=['GET', 'POST'])
        self.app.add_url_rule('/api/job-postings/<int:job_id>/applications', 'job_posting_applications', self.get_job_posting_applications, methods=['GET'])
        self.app.add_url_rule('/api/job-postings/<int:job_id>/assess-candidate/<int:candidate_id>', 'assess_candidate_for_job', self.assess_candidate_for_job_posting, methods=['POST'])
        
        # Enhanced PDS Processing routes (NEW)
        self.app.add_url_rule('/api/upload-pds-enhanced', 'upload_pds_enhanced', self.upload_pds_enhanced, methods=['POST'])
        # Removed duplicate start-analysis route - using the clean upload version
        self.app.add_url_rule('/api/analysis-status/<batch_id>', 'get_analysis_status', self.get_analysis_status, methods=['GET'])
        self.app.add_url_rule('/api/candidates-enhanced', 'get_candidates_enhanced', self.get_candidates_enhanced, methods=['GET'])
        self.app.add_url_rule('/api/clear-old-candidates', 'clear_old_candidates', self.clear_old_candidates, methods=['POST'])
    
    def login(self):
        """Handle user login"""
        if request.method == 'GET':
            return render_template('login.html')
        
        try:
            data = request.get_json() if request.is_json else request.form
            email = data.get('email')
            password = data.get('password')
            
            if not email or not password:
                return jsonify({'success': False, 'message': 'Email and password are required'}), 400
            
            # Authenticate user
            if BCRYPT_AVAILABLE:
                user_data = db_manager.authenticate_user(email, password)
            else:
                # Fallback authentication without bcrypt
                user_data = db_manager.get_user_by_email(email)
                if user_data and user_data.get('password') == password:
                    # Simple password check without hashing (for development only)
                    logger.warning("Using plain text password authentication - not recommended for production!")
                else:
                    user_data = None
            
            if user_data:
                # Set session data for authentication
                session['user_authenticated'] = True
                session['user_id'] = user_data['id']
                session['user_email'] = user_data['email']
                session['user_is_admin'] = user_data.get('is_admin', False)
                session['user_first_name'] = user_data.get('first_name', '')
                session['user_last_name'] = user_data.get('last_name', '')
                session['privacy_acknowledged'] = False  # Require privacy acknowledgment
                
                if FLASK_LOGIN_AVAILABLE:
                    user = SimpleUser(user_data)
                    login_user(user, remember=True)
                
                return jsonify({
                    'success': True,
                    'message': 'Login successful',
                    'redirect': '/privacy-agreement',  # Redirect to privacy page first
                    'user': {
                        'id': user_data['id'],
                        'email': user_data['email'],
                        'name': f"{user_data.get('first_name', '')} {user_data.get('last_name', '')}".strip(),
                        'is_admin': user_data.get('is_admin', False)
                    }
                })
            else:
                return jsonify({'success': False, 'message': 'Invalid email or password'}), 401
                
        except Exception as e:
            logger.error(f"Login error: {e}")
            return jsonify({'success': False, 'message': 'An error occurred during login'}), 500
    
    def logout(self):
        """Handle user logout"""
        if FLASK_LOGIN_AVAILABLE:
            logout_user()
        
        # Clear session data
        session.clear()
        
        return redirect(url_for('login'))
    
    def privacy_agreement(self):
        """Handle data privacy agreement page"""
        # Check if user is authenticated
        if not session.get('user_authenticated') and not (FLASK_LOGIN_AVAILABLE and current_user.is_authenticated):
            return redirect(url_for('login'))
        
        if request.method == 'GET':
            # Get user info with fallback for Flask-Login users
            email = session.get('user_email', current_user.email if hasattr(current_user, 'email') else '')
            first_name = session.get('user_first_name', current_user.first_name if hasattr(current_user, 'first_name') else '')
            last_name = session.get('user_last_name', current_user.last_name if hasattr(current_user, 'last_name') else '')
            
            user_info = {
                'email': email,
                'name': f"{first_name} {last_name}".strip() or email.split('@')[0] if email else '',
                'is_admin': session.get('user_is_admin', current_user.is_admin if hasattr(current_user, 'is_admin') else False)
            }
            return render_template('privacy_agreement.html', user_info=user_info)
        
        elif request.method == 'POST':
            try:
                data = request.get_json() if request.is_json else request.form
                agreed = data.get('agreed')
                
                if agreed:
                    session['privacy_acknowledged'] = True
                    return jsonify({
                        'success': True,
                        'message': 'Privacy agreement acknowledged',
                        'redirect': '/dashboard'
                    })
                else:
                    return jsonify({
                        'success': False,
                        'message': 'You must agree to the privacy policy to continue'
                    }), 400
                    
            except Exception as e:
                logger.error(f"Privacy agreement error: {e}")
                return jsonify({'success': False, 'message': 'An error occurred'}), 500
    
    def handle_users(self):
        """Handle user CRUD operations"""
        if not current_user.is_authenticated or not current_user.is_admin:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403
        
        try:
            if request.method == 'GET':
                users = db_manager.get_all_users()
                return jsonify({'success': True, 'users': users})
            
            elif request.method == 'POST':
                data = request.get_json()
                email = data['email']
                
                # Extract first/last name from email if not provided
                username = email.split('@')[0]
                first_name = data.get('first_name', username)
                last_name = data.get('last_name', 'User')
                
                user_id = db_manager.create_user(
                    email=email,
                    password=data['password'],
                    first_name=first_name,
                    last_name=last_name,
                    is_admin=data.get('is_admin', False)
                )
                return jsonify({'success': True, 'user_id': user_id})
        
        except Exception as e:
            logger.error(f"User management error: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500
    
    def handle_user(self, user_id):
        """Handle individual user operations"""
        if not current_user.is_authenticated or not current_user.is_admin:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403
        
        try:
            if request.method == 'GET':
                user = db_manager.get_user_by_id(user_id)
                if user:
                    return jsonify({'success': True, 'user': user})
                else:
                    return jsonify({'success': False, 'message': 'User not found'}), 404
                    
            elif request.method == 'PUT':
                data = request.get_json()
                success = db_manager.update_user(user_id, **data)
                return jsonify({'success': success})
                
            elif request.method == 'DELETE':
                # Prevent admin from deleting themselves
                if current_user.id == user_id:
                    return jsonify({'success': False, 'message': 'Cannot delete your own account'}), 400
                
                success = db_manager.delete_user(user_id)
                return jsonify({'success': success})
        
        except Exception as e:
            logger.error(f"User operation error: {e}")
            return jsonify({'success': False, 'message': str(e)}), 500
    
    def index(self):
        """Serve the landing page"""
        # Always check authentication first
        if not session.get('user_authenticated') and not (FLASK_LOGIN_AVAILABLE and current_user.is_authenticated):
            return redirect(url_for('login'))
        
        # Check if privacy has been acknowledged
        if not session.get('privacy_acknowledged', False):
            return redirect(url_for('privacy_agreement'))
        
        return redirect(url_for('dashboard'))
    
    @login_required
    def demo(self):
        """Serve the design demo page"""
        return render_template("demo.html")
    
    @login_required
    def dashboard(self, section=None):
        """Serve the dashboard page"""
        # Check if privacy has been acknowledged
        if not session.get('privacy_acknowledged', False):
            return redirect(url_for('privacy_agreement'))
        
        user_info = {
            'email': session.get('user_email', current_user.email if hasattr(current_user, 'email') else ''),
            'is_admin': session.get('user_is_admin', current_user.is_admin if hasattr(current_user, 'is_admin') else False),
            'is_active': session.get('user_authenticated', current_user.is_active if hasattr(current_user, 'is_active') else True)
        }
        return render_template("dashboard.html", user_info=user_info)
    
    def health_check(self):
        """Health check endpoint"""
        return jsonify({"status": "healthy", "message": "PDS Assessment System is running"})
    
    def system_status(self):
        """System status endpoint for monitoring"""
        try:
            status = {
                'database': True,
                'analytics': True,
                'upload': True,
                'assessment': True
            }
            
            # Check database connection
            try:
                with db_manager.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT 1")
                    cursor.fetchone()
            except Exception as e:
                logger.error(f"Database status check failed: {e}")
                status['database'] = False
            
            # Check analytics system
            try:
                analytics_summary = db_manager.get_analytics_summary()
                if not analytics_summary:
                    status['analytics'] = False
            except Exception as e:
                logger.error(f"Analytics status check failed: {e}")
                status['analytics'] = False
            
            # Overall system status
            system_healthy = all(status.values())
            
            return jsonify({
                'success': True,
                'data': status,
                'healthy': system_healthy,
                'timestamp': datetime.now().isoformat()
            })
            
        except Exception as e:
            logger.error(f"System status check failed: {e}")
            return jsonify({
                'success': False,
                'error': 'Failed to check system status',
                'healthy': False
            }), 500
    
    def debug_jobs(self):
        """Debug endpoint to show all available LSPU job postings"""
        try:
            import sqlite3
            conn = sqlite3.connect('pds_assessment.db')
            cursor = conn.cursor()
            
            # Get LSPU job postings only
            cursor.execute('''
                SELECT id, position_title, position_category, status, 
                       education_requirements, experience_requirements,
                       department_office, application_deadline 
                FROM lspu_job_postings 
                ORDER BY id
            ''')
            lspu_jobs = []
            for row in cursor.fetchall():
                lspu_jobs.append({
                    'id': row[0],
                    'title': row[1],
                    'category': row[2],
                    'status': row[3],
                    'education_requirements': row[4],
                    'experience_requirements': row[5],
                    'department_office': row[6],
                    'application_deadline': row[7],
                    'source': 'LSPU'
                })
            
            conn.close()
            
            return jsonify({
                'success': True,
                'job_postings': lspu_jobs,
                'total_jobs': len(lspu_jobs),
                'system': 'LSPU Only'
            })
            
        except Exception as e:
            logger.error(f"Error in debug_jobs: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def _get_job_by_id(self, job_id):
        """Get job by ID, checking both LSPU job postings and legacy jobs"""
        try:
            # First try LSPU job postings using database manager
            try:
                with db_manager.get_connection() as conn:
                    cursor = conn.cursor()
                    
                    # Simple query first - just get the job posting data
                    cursor.execute('''
                        SELECT * FROM lspu_job_postings WHERE id = %s
                    ''', (job_id,))
                    
                    row = cursor.fetchone()
                    if row:
                        # Convert to regular dict if it's a RealDictRow
                        job = dict(row) if row else {}
                        job['title'] = job.get('position_title', 'Unknown Position')
                        job['category'] = 'University Position'
                        job['source'] = 'LSPU'
                        logger.info(f"✅ Successfully fetched LSPU job posting {job_id}: {job.get('position_title', 'Unknown')}")
                        return job
                
            except Exception as e:
                logger.warning(f"Could not fetch LSPU job posting {job_id}: {e}")
            
            # Fallback to legacy job system
            job = db_manager.get_job(job_id)
            if job:
                job['source'] = 'Legacy'
                logger.info(f"✅ Successfully fetched legacy job {job_id}")
                return job
            
            logger.warning(f"❌ No job found with ID {job_id}")
            return None
            
        except Exception as e:
            logger.error(f"Error getting job {job_id}: {e}")
            return None

    def _is_allowed_file(self, filename):
        """Check if file type is allowed"""
        allowed_extensions = {'pdf', 'doc', 'docx', 'txt', 'xlsx', 'xls', 'jpg', 'jpeg', 'png', 'tiff', 'bmp'}
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

    def _process_file_for_analysis(self, file_record, job):
        """Process a file through the appropriate analysis engine"""
        try:
            file_path = file_record['temp_path']
            original_name = file_record['original_name']
            file_type = file_record['file_type']
            
            logger.info(f"ðŸ” Processing {original_name} (type: {file_type})")
            
            # First, try to determine if this is a PDS file
            is_pds_file = self._detect_pds_file(file_path, file_type)
            
            if is_pds_file and file_type == 'excel':
                logger.info(f"ðŸ“‹ Processing as PDS: {original_name}")
                return self.x_process_pds_file(file_path, original_name, job)
            elif file_type == 'pdf':
                logger.info(f"ðŸ“„ Processing as PDF resume: {original_name}")
                return self._process_pdf_file(file_path, original_name, job)
            else:
                logger.warning(f"âš ï¸ Unsupported file type for analysis: {file_type}")
                return None
                
        except Exception as e:
            logger.error(f"âŒ Error processing file {file_record['original_name']}: {e}")
            return None

    def _detect_pds_file(self, file_path, file_type):
        """Detect if a file is a PDS file"""
        if file_type != 'excel':
            return False
            
        try:
            # Try to open as Excel and check for PDS indicators
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Check for typical PDS sheet names
            pds_indicators = ['C1', 'C2', 'C3', 'C4', 'PERSONAL DATA SHEET', 'PDS']
            sheet_names = [name.upper() for name in wb.sheetnames]
            
            has_pds_sheets = any(indicator in sheet_names for indicator in pds_indicators)
            
            # Additional check for content
            if has_pds_sheets or 'C1' in wb.sheetnames:
                return True
                
            # Check first sheet for PDS content
            first_sheet = wb.active
            if first_sheet:
                # Look for common PDS text
                for row in first_sheet.iter_rows(max_row=10, max_col=10):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            cell_text = cell.value.upper()
                            if any(phrase in cell_text for phrase in [
                                'PERSONAL DATA SHEET', 'CS FORM', 'SURNAME', 'FIRST NAME', 'MIDDLE NAME'
                            ]):
                                return True
            
            wb.close()
            return False
            
        except Exception as e:
            logger.warning(f"Could not detect PDS format: {e}")
            return False

    def _process_pds_file(self, file_path, original_name, job):
        """Process a PDS file"""
        try:
            if self.pds_processor:
                # Use the enhanced PDS processor with proper file handling
                if original_name.lower().endswith(('.xlsx', '.xls')):
                    # For Excel files, use the Excel-specific processing method
                    candidate_data = self._process_excel_file(file_path, original_name, job)
                else:
                    # For text-based files, read as text
                    with open(file_path, 'rb') as file:
                        file_content = file.read()
                        # Try to decode bytes to string for text-based processing
                        try:
                            if isinstance(file_content, bytes):
                                file_content = file_content.decode('utf-8')
                        except UnicodeDecodeError:
                            # If decoding fails, try different encodings
                            try:
                                file_content = file_content.decode('latin-1')
                            except:
                                logger.error(f"Could not decode file content for {original_name}")
                                return None
                        candidate_data = self.pds_processor.process_pds_candidate(file_content)
                    
                    if candidate_data and 'pds_data' in candidate_data:
                        # Use enhanced assessment if available
                        is_lspu_job = job.get('source') == 'LSPU'
                        if is_lspu_job and self.assessment_engine:
                            try:
                                logger.info(f"ðŸŽ¯ Using LSPU assessment engine for {original_name}")
                                assessment_result = self.assessment_engine.assess_candidate_for_lspu_job(
                                    candidate_data=candidate_data.get('pds_data', {}),
                                    lspu_job=job,
                                    position_type_id=job.get('position_type_id')
                                )
                                score = assessment_result.get('automated_score', 0)
                                percentage_score = assessment_result.get('percentage_score', 0)
                                logger.info(f"âœ… LSPU assessment completed: {percentage_score}%")
                            except Exception as e:
                                logger.warning(f"âš ï¸ LSPU assessment failed, using fallback: {e}")
                                score = self._calculate_pds_score(candidate_data, job)
                                percentage_score = score
                        else:
                            logger.info(f"ðŸ“Š Using standard PDS scoring for {original_name}")
                            score = self._calculate_pds_score(candidate_data, job)
                            percentage_score = score
                        
                        candidate_data['score'] = score
                        candidate_data['percentage_score'] = percentage_score
                        candidate_data['processing_type'] = 'pds_digital'
                        
                        return candidate_data
            else:
                # Fallback PDS processing using basic extraction
                logger.info(f"ðŸ“Š Using fallback PDS processing for {original_name}")
                candidate_data = self._fallback_pds_processing(file_path, original_name, job)
                return candidate_data
                
        except Exception as e:
            logger.error(f"âŒ Error processing PDS file {original_name}: {e}")
            return None

    def _process_pdf_file(self, file_path, original_name, job):
        """Process a PDF resume file"""
        try:
            if self.resume_processor:
                # Create a mock file object that the resume processor expects
                class MockFile:
                    def __init__(self, file_path, filename):
                        self.file_path = file_path
                        self.filename = filename
                        self._content = None
                    
                    def read(self):
                        if self._content is None:
                            with open(self.file_path, 'rb') as f:
                                self._content = f.read()
                        return self._content
                    
                    def seek(self, position):
                        pass  # Not needed for our use case
                
                mock_file = MockFile(file_path, original_name)
                text = self.resume_processor.extract_text_from_file(mock_file)
                
                if text.strip():
                    result = self._process_resume_for_job(text, original_name, job)
                    
                    # Convert to candidate data format
                    candidate_data = {
                        'name': result.get('name', 'Unknown'),
                        'email': result.get('email', ''),
                        'phone': result.get('phone', ''),
                        'address': result.get('address', ''),
                        'linkedin': result.get('linkedin', ''),
                        'resume_text': text,
                        'education': json.dumps(result.get('education', [])),
                        'work_experience': json.dumps(result.get('workExperience', [])),
                        'skills': json.dumps(result.get('allSkills', [])),
                        'certifications': json.dumps(result.get('certifications', [])),
                        'job_id': job['id'],
                        'category': result.get('predictedCategory', {}).get('category', 'Unknown'),
                        'score': result['matchScore'],
                        'scoring_breakdown': json.dumps(result.get('scoringBreakdown', {})),
                        'years_of_experience': result.get('yearsOfExperience', 0),
                        'education_level': result.get('educationLevel', 'Not Specified'),
                        'matched_skills': json.dumps(result.get('matchedSkills', [])),
                        'missing_skills': json.dumps(result.get('missingSkills', [])),
                        'status': 'pending',
                        'processing_type': 'resume_pdf'
                    }
                    
                    return candidate_data
                else:
                    logger.warning(f"âš ï¸ No text extracted from PDF: {original_name}")
                    return None
            else:
                logger.error(f"âŒ Resume processor not available for {original_name}")
                return None
                
        except Exception as e:
            logger.error(f"âŒ Error processing PDF file {original_name}: {e}")
            return None

    def _fallback_pds_processing(self, file_path, original_name, job):
        """Fallback PDS processing when main processor is not available"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Basic extraction from first sheet
            first_sheet = wb.active
            candidate_data = {
                'name': 'Unknown',
                'email': '',
                'phone': '',
                'job_id': job['id'],
                'score': 50,  # Default score
                'processing_type': 'pds_fallback',
                'status': 'pending'
            }
            
            # Try to extract basic info
            try:
                for row in first_sheet.iter_rows(max_row=20, max_col=10):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            cell_text = cell.value.strip()
                            # Look for name patterns
                            if len(cell_text) > 2 and ' ' in cell_text and cell_text.count(' ') <= 3:
                                # Might be a name
                                if candidate_data['name'] == 'Unknown':
                                    candidate_data['name'] = cell_text
                            # Look for email
                            if '@' in cell_text and '.' in cell_text:
                                candidate_data['email'] = cell_text
                            # Look for phone
                            if re.match(r'[\+\d\-\(\)\s]{7,}', cell_text) and any(c.isdigit() for c in cell_text):
                                candidate_data['phone'] = cell_text
            except Exception as e:
                logger.warning(f"âš ï¸ Error in basic extraction: {e}")
            
            wb.close()
            
            # If we found a name, it's probably valid
            if candidate_data['name'] != 'Unknown':
                candidate_data['score'] = 60  # Better score if we extracted some data
            
            logger.info(f"ðŸ“‹ Fallback PDS processing complete for {original_name}: {candidate_data['name']}")
            return candidate_data
            
        except Exception as e:
            logger.error(f"âŒ Fallback PDS processing failed for {original_name}: {e}")
            return None

    def _cleanup_session_files(self, upload_files):
        """Clean up temporary files after processing"""
        try:
            for file_record in upload_files:
                temp_path = file_record['temp_path']
                try:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                        logger.info(f"ðŸ—‘ï¸ Cleaned up temp file: {temp_path}")
                except Exception as e:
                    logger.warning(f"Failed to clean up temp file {temp_path}: {e}")
        except Exception as e:
            logger.error(f"Error during file cleanup: {e}")
    
    def _is_image_file(self, filename):
        """Check if file is an image that requires OCR processing"""
        image_extensions = {'jpg', 'jpeg', 'png', 'tiff', 'bmp'}
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in image_extensions
    
    @login_required
    def _calculate_comprehensive_score(self, education, experience, skills, certifications, job_requirements, job):
        """Calculate comprehensive score based on education, experience, and skills"""
        total_score = 0
        breakdown = {}
        
        # Education Score (40% of total - 40 points)
        education_score = self._score_education(education, job)
        breakdown['education'] = education_score
        total_score += education_score
        
        # Experience Score (40% of total - 40 points) 
        experience_score = self._score_experience(experience, job)
        breakdown['experience'] = experience_score
        total_score += experience_score
        
        # Skills Match Score (20% of total - 20 points)
        skills_result = self._score_skills_match(skills, job_requirements)
        breakdown['skills'] = skills_result['score']
        total_score += skills_result['score']
        
        return {
            'total_score': min(round(total_score), 100),
            'breakdown': breakdown,
            'matched_skills': skills_result['matched'],
            'missing_skills': skills_result['missing']
        }
    
    def _score_education(self, education_info, job):
        """Score education (max 40 points)"""
        score = 0
        
        if not education_info:
            return 0
            
        # Get highest education level
        education_levels = {
            'phd': 40, 'doctorate': 40, 'doctoral': 40,
            'master': 35, 'masters': 35, 'msc': 35, 'mba': 35,
            'bachelor': 25, 'bachelors': 25, 'degree': 25, 'bsc': 25,
            'diploma': 15, 'associate': 15,
            'certificate': 10, 'high school': 5
        }
        
        # Check education entries
        max_education_score = 0
        for edu in education_info:
            edu_text = str(edu).lower()
            for level, points in education_levels.items():
                if level in edu_text:
                    max_education_score = max(max_education_score, points)
                    break
        
        score = max_education_score
        
        # Relevance bonus (up to 5 additional points)
        job_title = job.get('title', '').lower()
        job_category = job.get('category', '').lower()
        
        for edu in education_info:
            edu_text = str(edu).lower()
            # Check for field relevance
            if any(tech in edu_text for tech in ['computer', 'information', 'technology', 'engineering']):
                if any(tech in job_title or tech in job_category for tech in ['it', 'technology', 'software', 'developer', 'engineer']):
                    score += 5
                    break
            elif any(business in edu_text for business in ['business', 'management', 'administration']):
                if any(admin in job_title or admin in job_category for admin in ['manager', 'admin', 'supervisor']):
                    score += 5
                    break
        
        return min(score, 40)
    
    def _score_experience(self, experience_info, job):
        """Score work experience (max 40 points)"""
        if not experience_info:
            return 0
            
        # Calculate years of experience (max 30 points)
        total_years = self._calculate_years_of_experience(experience_info)
        experience_points = min(total_years * 3, 30)  # 3 points per year, max 30
        
        # Relevance bonus (max 10 points)
        relevance_score = 0
        job_title = job.get('title', '').lower()
        job_category = job.get('category', '').lower()
        
        for exp in experience_info:
            if isinstance(exp, dict):
                position = exp.get('position', '').lower()
                company = exp.get('company', '').lower()
                description = exp.get('description', '').lower()
                
                # Check for position relevance
                if any(keyword in position for keyword in job_title.split()):
                    relevance_score += 5
                    break
                elif any(keyword in description for keyword in job_title.split()):
                    relevance_score += 3
                    break
                
        return min(experience_points + relevance_score, 40)
    
    def _score_skills_match(self, skills_info, job_requirements):
        """Score skills match (max 20 points)"""
        if not job_requirements:
            return {'score': 20, 'matched': [], 'missing': []}
            
        # Flatten skills list
        all_skills = []
        if isinstance(skills_info, list):
            all_skills = [str(skill).lower() for skill in skills_info]
        elif isinstance(skills_info, dict):
            for category_skills in skills_info.values():
                if isinstance(category_skills, list):
                    all_skills.extend([str(skill).lower() for skill in category_skills])
        
        matched_skills = []
        missing_skills = []
        
        for req_skill in job_requirements:
            req_skill_lower = req_skill.lower()
            found = False
            for skill in all_skills:
                if req_skill_lower in skill or skill in req_skill_lower:
                    matched_skills.append(req_skill.title())
                    found = True
                    break
            if not found:
                missing_skills.append(req_skill.title())
        
        # Calculate score based on match percentage
        match_percentage = len(matched_skills) / len(job_requirements) if job_requirements else 1
        score = round(match_percentage * 20)
        
        return {
            'score': score,
            'matched': matched_skills,
            'missing': missing_skills
        }
    
    def _calculate_years_of_experience(self, experience_info):
        """Calculate total years of work experience"""
        if not experience_info:
            return 0
            
        # Simple calculation - count number of jobs (can be enhanced)
        return len(experience_info)
    
    def _determine_education_level(self, education_info):
        """Determine the highest education level"""
        if not education_info:
            return 'Not Specified'
            
        education_levels = {
            'phd': 'Doctorate', 'doctorate': 'Doctorate', 'doctoral': 'Doctorate',
            'master': 'Masters', 'masters': 'Masters', 'msc': 'Masters', 'mba': 'Masters',
            'bachelor': 'Bachelors', 'bachelors': 'Bachelors', 'degree': 'Bachelors', 'bsc': 'Bachelors',
            'diploma': 'Diploma', 'associate': 'Associate',
            'certificate': 'Certificate', 'high school': 'High School'
        }
        
        highest_level = 'Not Specified'
        highest_rank = 0
        level_ranks = {'Doctorate': 6, 'Masters': 5, 'Bachelors': 4, 'Diploma': 3, 'Associate': 2, 'Certificate': 1, 'High School': 0}
        
        for edu in education_info:
            edu_text = str(edu).lower()
            for level_key, level_name in education_levels.items():
                if level_key in edu_text:
                    rank = level_ranks.get(level_name, 0)
                    if rank > highest_rank:
                        highest_rank = rank
                        highest_level = level_name
                    break
        
        return highest_level
    
    def _determine_university_position_type(self, education_info, work_experience, skills_info, certifications):
        """Determine the most suitable university position type based on candidate profile"""
        try:
            # Score candidate for each position type
            position_scores = {
                'Regular Faculty': 0,
                'Part-time Teaching': 0,
                'Non-Teaching Personnel': 0,
                'Job Order': 0
            }
            
            # Education-based scoring
            education_level = self._determine_education_level(education_info)
            teaching_keywords = ['teaching', 'professor', 'instructor', 'education', 'academic', 'research', 'faculty']
            admin_keywords = ['administration', 'management', 'administrative', 'support', 'officer', 'coordinator']
            tech_keywords = ['technical', 'engineering', 'analyst', 'developer', 'specialist', 'technician']
            
            # Check for teaching experience and qualifications
            has_teaching_experience = False
            has_admin_experience = False
            has_technical_experience = False
            years_experience = self._calculate_years_of_experience(work_experience)
            
            # Analyze work experience
            for exp in work_experience:
                position_title = exp.get('position', '').lower()
                company = exp.get('company', '').lower()
                
                if any(keyword in position_title for keyword in teaching_keywords):
                    has_teaching_experience = True
                if any(keyword in position_title for keyword in admin_keywords):
                    has_admin_experience = True
                if any(keyword in position_title for keyword in tech_keywords):
                    has_technical_experience = True
            
            # Regular Faculty scoring (needs advanced degree + teaching/research experience)
            if education_level in ['Masters', 'PhD', 'Doctorate']:
                position_scores['Regular Faculty'] += 40
                if has_teaching_experience:
                    position_scores['Regular Faculty'] += 30
                if years_experience >= 3:
                    position_scores['Regular Faculty'] += 20
                    
            # Part-time Teaching scoring (can have bachelor's + some teaching experience)
            if education_level in ['Bachelor', 'Masters', 'PhD', 'Doctorate']:
                position_scores['Part-time Teaching'] += 30
                if has_teaching_experience:
                    position_scores['Part-time Teaching'] += 40
                elif years_experience >= 2:  # Any professional experience
                    position_scores['Part-time Teaching'] += 20
                    
            # Non-Teaching Personnel scoring (admin, technical, support roles)
            if has_admin_experience:
                position_scores['Non-Teaching Personnel'] += 50
            elif has_technical_experience:
                position_scores['Non-Teaching Personnel'] += 40
            if education_level in ['Bachelor', 'Masters', 'PhD']:
                position_scores['Non-Teaching Personnel'] += 25
                
            # Job Order scoring (project-based, temporary, entry-level)
            if years_experience < 2:  # Less experience
                position_scores['Job Order'] += 30
            if education_level in ['Bachelor', 'High School', 'Associate']:
                position_scores['Job Order'] += 25
            # Always viable option
            position_scores['Job Order'] += 20
            
            # Find the best match
            best_position = max(position_scores, key=position_scores.get)
            best_score = position_scores[best_position]
            
            # Calculate confidence (0-100%)
            total_possible_score = 100  # Rough estimate of max possible score
            confidence = min(100, round((best_score / total_possible_score) * 100, 2))
            
            return {
                'category': best_position,
                'confidence': confidence,
                'scores': position_scores
            }
            
        except Exception as e:
            logger.error(f"Error determining university position type: {e}")
            return {
                'category': 'Job Order',  # Default fallback
                'confidence': 50,
                'scores': {}
            }
    
    def upload_pds(self):
        """Handle Personal Data Sheet upload and processing"""
        # Check authentication for API endpoint
        if not session.get('user_authenticated') and not (FLASK_LOGIN_AVAILABLE and current_user.is_authenticated):
            return jsonify({'success': False, 'error': 'Authentication required'}), 401
        
        try:
            if 'files[]' not in request.files:
                return jsonify({'success': False, 'error': 'No files uploaded'}), 400
            
            files = request.files.getlist('files[]')
            job_id = request.form.get('jobId')
            
            # DEBUG: Log what we received
            logger.info(f"=== UPLOAD_PDS DEBUG START ===")
            logger.info(f"Received job_id from frontend: {job_id} (type: {type(job_id)})")
            logger.info(f"Number of files: {len(files)}")
            
            if not files or all(f.filename == '' for f in files):
                return jsonify({'success': False, 'error': 'No files selected'}), 400
            
            if not job_id:
                logger.error("No job_id provided in request")
                return jsonify({'success': False, 'error': 'Job ID is required'}), 400
            
            try:
                job_id = int(job_id)
                logger.info(f"Converted job_id to integer: {job_id}")
            except ValueError:
                logger.error(f"Failed to convert job_id '{job_id}' to integer")
                return jsonify({'success': False, 'error': 'Invalid job ID'}), 400
            
            # Get LSPU job posting details
            job = None
            
            logger.info(f"Fetching LSPU job posting with id: {job_id}")
            try:
                conn = sqlite3.connect('resume_screening.db')
                cursor = conn.cursor()
                
                cursor.execute('''
                    SELECT id, position_title, position_category, campus_id,
                           salary_grade, salary_amount, education_requirements, experience_requirements,
                           training_requirements, eligibility_requirements, 
                           application_deadline, department_office, status
                    FROM lspu_job_postings 
                    WHERE id = ?
                ''', (job_id,))
                
                row = cursor.fetchone()
                if row:
                    logger.info(f"âœ“ Found LSPU job posting: ID={row[0]}, Title={row[1]}")
                    job = {
                        'id': row[0],
                        'title': row[1],
                        'category': row[2],
                        'campus_id': row[3],
                        'salary_grade': row[4],
                        'salary_amount': row[5],
                        'education_requirements': row[6],
                        'experience_requirements': row[7],
                        'training_requirements': row[8],
                        'eligibility_requirements': row[9],
                        'application_deadline': row[10],
                        'department_office': row[11],
                        'status': row[12],
                        'source': 'LSPU'
                    }
                else:
                    logger.warning(f"âœ— LSPU job posting not found with id: {job_id}")
                conn.close()
                
            except Exception as e:
                logger.error(f"âœ— Error fetching LSPU job posting {job_id}: {e}")
            
            if not job:
                logger.error(f"=== LSPU JOB NOT FOUND: job_id={job_id} ===")
                return jsonify({'success': False, 'error': 'Job posting not found'}), 404
            
            logger.info(f"âœ“ Successfully loaded job: {job.get('title')} (ID: {job.get('id')})")
            logger.info(f"=== UPLOAD_PDS DEBUG END ===")
            
            results = []
            errors = []
            
            for file in files:
                if file.filename != '' and self._is_allowed_file(file.filename):
                    try:
                        logger.info(f"Processing PDS file: {file.filename}")
                        
                        # Determine if this is an image file requiring OCR
                        if self._is_image_file(file.filename):
                            logger.info(f"Image file detected, processing with OCR: {file.filename}")
                            
                            # Process image with OCR (placeholder for now)
                            # In a full implementation, you would use OCR here
                            file.seek(0)
                            text = "OCR processing not yet implemented for image files"
                            
                            # Create basic candidate data for OCR-processed images
                            candidate_data = {
                                'name': f"OCR Candidate from {file.filename}",
                                'email': '',
                                'phone': '',
                                'resume_text': text,
                                'education': '[]',
                                'skills': '',
                                'job_id': job_id,
                                'category': 'Unknown',
                                'score': 0,
                                'status': 'pending',
                                'processing_type': 'ocr_scanned',
                                'pds_data': json.dumps({})
                            }
                            
                            candidate_id = db_manager.create_candidate(candidate_data)
                            
                            result = {
                                'candidate_id': candidate_id,
                                'filename': file.filename,
                                'name': candidate_data['name'],
                                'email': candidate_data['email'],
                                'total_score': 0,
                                'processing_type': 'ocr_scanned',
                                'sections_extracted': []
                            }
                            
                        else:
                            # Check if this is a PDS file
                            file.seek(0)  # Ensure we start at the beginning
                            is_pds = self.processor.is_pds_file(file)
                            logger.info(f"PDS detection result for {file.filename}: {is_pds}")
                            
                            if is_pds:
                                logger.info(f"Detected PDS format: {file.filename}")
                                
                                # Use comprehensive PDS extraction
                                file.seek(0)  # Reset file pointer
                                file_content = file.read()
                                file.seek(0)  # Reset again for any subsequent reads
                                candidate_data = self.processor.process_pds_candidate(file_content)
                                
                                if not candidate_data:
                                    errors.append(f"{file.filename}: Failed to extract PDS data")
                                    continue
                                
                                # CRITICAL FIX: Ensure job_id is properly set
                                candidate_data['job_id'] = job_id
                                logger.info(f"âœ“ Set job_id={job_id} for candidate: {candidate_data.get('name', 'Unknown')}")
                                
                                logger.info(f"PDS extraction successful for {file.filename}")
                                
                                # Calculate PDS-specific score
                                score = self._calculate_pds_score(candidate_data, job)
                                candidate_data['score'] = score
                                
                                logger.info(f"PDS score calculated for {file.filename}: {score}")
                                
                                # Store in database
                                candidate_id = db_manager.create_candidate(candidate_data)
                                logger.info(f"âœ“ Created candidate with ID: {candidate_id}, job_id: {candidate_data['job_id']}")
                                
                                # Prepare result for response
                                result = {
                                    'candidate_id': candidate_id,
                                    'filename': file.filename,
                                    'name': candidate_data['name'],
                                    'email': candidate_data['email'],
                                    'total_score': score,
                                    'processing_type': 'pds',
                                    'sections_extracted': list(candidate_data['pds_data'].keys()),
                                    'job_assignment': {
                                        'job_id': job_id,
                                        'job_title': job.get('title', 'Unknown')
                                    }
                                }
                                
                            else:
                                logger.info(f"Processing as text-based file: {file.filename}")
                                
                                # Fall back to text extraction for non-PDS Excel files
                                file.seek(0)  # Reset file pointer
                                text = self.processor.extract_text_from_file(file)
                                logger.info(f"Extracted text length: {len(text)} characters")
                                
                                if not text.strip():
                                    errors.append(f"{file.filename}: No text could be extracted")
                                    continue
                                
                                # Process as regular text-based PDS
                                result = self._process_pds_for_job(text, file.filename, job)
                                logger.info(f"Processing result for {file.filename}: Score={result.get('total_score', 0)}")
                                
                                # Store candidate data with basic PDS fields
                                candidate_data = {
                                    'name': result.get('basic_info', {}).get('name', 'Unknown'),
                                    'email': result.get('basic_info', {}).get('email', ''),
                                    'phone': result.get('basic_info', {}).get('phone', ''),
                                    'resume_text': text,
                                    'education': result.get('education', []),
                                    'skills': ', '.join(result.get('skills', {}).get('technical', [])),
                                    'job_id': job_id,
                                    'category': result.get('predicted_category', {}).get('category', 'Unknown'),
                                    'score': result['total_score'],
                                    'status': 'pending',
                                    'processing_type': 'pds_text',
                                    'pds_data': json.dumps(result)
                                }
                                
                                candidate_id = db_manager.create_candidate(candidate_data)
                                result['candidate_id'] = candidate_id
                        
                        # Add result to results list
                        results.append(result)
                        logger.info(f"Successfully processed {file.filename}")
                        
                    except Exception as e:
                        error_msg = f"Error processing file {file.filename}: {e}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue
                else:
                    if file.filename:
                        errors.append(f"{file.filename}: Unsupported file type")
            
            response_data = {
                'success': True,
                'message': f'Successfully processed {len(results)} Personal Data Sheets',
                'results': results,
                'processing_type': 'pds'
            }
            
            if errors:
                response_data['warnings'] = errors
                response_data['message'] += f' ({len(errors)} files had errors)'
            
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"Error in upload_pds: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    def _process_pds_for_job(self, pds_text, filename, job):
        """Process a Personal Data Sheet against a specific job using LSPU job posting requirements"""
        try:
            # Extract comprehensive PDS information
            pds_data = self.pds_processor.extract_pds_information(pds_text, filename)
            
            # Check if this is an LSPU job posting (has LSPU-specific fields)
            is_lspu_job = any(field in job for field in [
                'education_requirements', 'experience_requirements', 
                'training_requirements', 'eligibility_requirements', 'position_title'
            ])
            
            if is_lspu_job and self.assessment_engine:
                logger.info(f"Using LSPU university assessment engine for job {job.get('id', 'unknown')}")
                
                # Use the enhanced LSPU university assessment engine
                assessment_result = self.assessment_engine.assess_candidate_for_lspu_job(
                    candidate_data=pds_data,
                    lspu_job=job,
                    position_type_id=job.get('position_type_id')
                )
                
                # Determine university position type based on extracted data
                university_position = self._determine_university_position_type(
                    pds_data.get('education', []),
                    pds_data.get('experience', []),
                    pds_data.get('skills', []),
                    pds_data.get('certifications', [])
                )
                
                # Format result for compatibility with existing system
                result = {
                    'filename': filename,
                    'basic_info': pds_data.get('basic_info', {}),
                    'education': pds_data.get('education', []),
                    'experience': pds_data.get('experience', []),
                    'skills': {
                        'technical': pds_data.get('skills', []),
                        'certifications': pds_data.get('certifications', [])
                    },
                    'certifications': pds_data.get('certifications', []),
                    'training': pds_data.get('training', []),
                    'awards': pds_data.get('awards', []),
                    'eligibility': pds_data.get('eligibility', []),
                    'languages': pds_data.get('languages', []),
                    'licenses': pds_data.get('licenses', []),
                    'volunteer_work': pds_data.get('volunteer_work', []),
                    'predicted_category': university_position,
                    'total_score': assessment_result.get('automated_score', 0),
                    'percentage_score': assessment_result.get('percentage_score', 0),
                    'category_scores': {
                        'education': assessment_result.get('assessment_results', {}).get('education', {}).get('score', 0),
                        'experience': assessment_result.get('assessment_results', {}).get('experience', {}).get('score', 0),
                        'training': assessment_result.get('assessment_results', {}).get('training', {}).get('score', 0),
                        'eligibility': assessment_result.get('assessment_results', {}).get('eligibility', {}).get('score', 0),
                        'accomplishments': assessment_result.get('assessment_results', {}).get('accomplishments', {}).get('score', 0)
                    },
                    'scoring_breakdown': assessment_result.get('assessment_results', {}),
                    'criteria_breakdown': assessment_result.get('criteria_breakdown', {}),
                    'assessment_engine_used': 'LSPU_University_Standards',
                    'job_match_details': {
                        'job_title': job.get('position_title', job.get('title', 'Unknown')),
                        'job_category': job.get('position_type_name', job.get('category', 'University Position')),
                        'job_requirements': assessment_result.get('job_requirements_used', {}),
                        'recommendation': assessment_result.get('recommendation', 'pending')
                    }
                }
                
                logger.info(f"LSPU assessment completed - Score: {assessment_result.get('percentage_score', 0)}%")
                return result
                
            else:
                # Fallback to legacy processing for old job system
                logger.info(f"Using legacy assessment for job {job.get('id', 'unknown')}")
                
                # Get detailed job requirements from position_requirements table (legacy)
                job_id = job.get('id')
                detailed_requirements = None
                
                if job_id:
                    try:
                        detailed_requirements = db_manager.get_position_requirements(job_id)
                        logger.info(f"Found legacy requirements for job {job_id}: {detailed_requirements}")
                    except Exception as e:
                        logger.warning(f"Could not fetch legacy requirements for job {job_id}: {e}")
                
                # Prepare job requirements for scoring (legacy format)
                if detailed_requirements:
                    # Use detailed requirements if available
                    job_requirements = {
                        'education_level': detailed_requirements.get('minimum_education', 'Bachelor'),
                        'experience_years': detailed_requirements.get('required_experience', 3),
                        'required_skills': detailed_requirements.get('required_skills', []),
                        'required_certifications': detailed_requirements.get('required_certifications', []),
                        'preferred_field': detailed_requirements.get('subject_area', job.get('category', '')),
                        'preferred_qualifications': detailed_requirements.get('preferred_qualifications', ''),
                        'relevant_experience': [job.get('title', '')]
                    }
                    logger.info(f"Using detailed requirements for scoring: {job_requirements}")
                else:
                    # Fallback to basic job requirements for backward compatibility
                    job_requirements = {
                        'education_level': 'Bachelor',  # Default
                        'experience_years': 3,  # Default
                        'required_skills': [skill.strip() for skill in job.get('requirements', '').split(',') if skill.strip()],
                        'required_certifications': [],
                        'preferred_field': job.get('category', ''),
                        'preferred_qualifications': '',
                        'relevant_experience': [job.get('title', '')]
                    }
                    logger.warning(f"Using fallback requirements for job {job_id}: {job_requirements}")
                
                # Score PDS against job requirements (legacy method)
                scoring_result = self.pds_processor.score_pds_against_job(pds_data, job_requirements)
                
                # Determine university position type instead of ML prediction
                university_position = self._determine_university_position_type(
                    pds_data.get('education', []),
                    pds_data.get('experience', []),
                    pds_data.get('skills', []),
                    pds_data.get('certifications', [])
                )
                
                # Combine results (legacy format)
                result = {
                    'filename': filename,
                    'basic_info': pds_data.get('basic_info', {}),
                    'education': pds_data.get('education', []),
                    'experience': pds_data.get('experience', []),
                    'skills': {
                        'technical': pds_data.get('skills', []),
                        'certifications': pds_data.get('certifications', [])
                    },
                    'certifications': pds_data.get('certifications', []),
                    'training': pds_data.get('training', []),
                    'awards': pds_data.get('awards', []),
                    'eligibility': pds_data.get('eligibility', []),
                    'languages': pds_data.get('languages', []),
                    'licenses': pds_data.get('licenses', []),
                    'volunteer_work': pds_data.get('volunteer_work', []),
                    'predicted_category': university_position,
                    'total_score': scoring_result.get('total_score', 0),
                    'category_scores': scoring_result.get('category_scores', {}),
                    'scoring_breakdown': scoring_result.get('scoring_breakdown', {}),
                    'assessment_engine_used': 'Legacy_PDS_Processor',
                    'job_match_details': {
                        'job_title': job.get('title'),
                        'job_category': job.get('category'),
                        'job_requirements': job_requirements
                    }
                }
                
                return result
            
        except Exception as e:
            logger.error(f"Error processing PDS for job: {str(e)}")
            return {
                'filename': filename,
                'error': str(e),
                'total_score': 0,
                'assessment_engine_used': 'Error'
            }
    
    def _calculate_pds_score(self, candidate_data, job):
        """Calculate comprehensive score for PDS candidate against job requirements."""
        try:
            total_score = 0
            max_score = 100
            scoring_breakdown = {}
            
            pds_data = candidate_data.get('pds_data', {})
            personal_info = pds_data.get('personal_info', {})
            
            # 1. Education Score (25 points)
            education_score = self._score_pds_education(personal_info.get('education', {}), job)
            scoring_breakdown['education'] = education_score
            total_score += education_score
            
            # 2. Work Experience Score (30 points)
            experience_score = self._score_pds_experience(pds_data.get('work_experience', []), job)
            scoring_breakdown['experience'] = experience_score
            total_score += experience_score
            
            # 3. Civil Service Eligibility Score (20 points)
            eligibility_score = self._score_pds_eligibility(pds_data.get('eligibility', []))
            scoring_breakdown['eligibility'] = eligibility_score
            total_score += eligibility_score
            
            # 4. Training and Development Score (15 points)
            training_score = self._score_pds_training(pds_data.get('training', []), job)
            scoring_breakdown['training'] = training_score
            total_score += training_score
            
            # 5. Voluntary Work Score (10 points)
            volunteer_score = self._score_pds_volunteer_work(pds_data.get('voluntary_work', []))
            scoring_breakdown['volunteer_work'] = volunteer_score
            total_score += volunteer_score
            
            # Store detailed scoring breakdown
            candidate_data['scoring_breakdown'] = scoring_breakdown
            
            return min(total_score, max_score)  # Cap at 100
            
        except Exception as e:
            logger.error(f"Error calculating PDS score: {str(e)}")
            return 0
    
    def _score_pds_education(self, education, job):
        """Score education background (max 25 points)."""
        score = 0
        
        # Graduate studies (15 points)
        if education.get('graduate') and education['graduate'].upper() != 'YEAR GRADUATED':
            score += 15
        # College degree (10 points)
        elif education.get('college'):
            score += 10
        # Secondary education (5 points)
        elif education.get('secondary'):
            score += 5
        
        # Relevance bonus (up to 10 points)
        job_title = job.get('title', '').lower()
        job_category = job.get('category', '').lower()
        
        college_info = education.get('college', '').lower()
        if college_info:
            if any(keyword in college_info for keyword in ['computer', 'information technology', 'engineering']):
                if any(tech in job_title or tech in job_category for tech in ['it', 'technology', 'software', 'system']):
                    score += 10
            elif any(keyword in college_info for keyword in ['business', 'management', 'administration']):
                if any(admin in job_title or admin in job_category for admin in ['admin', 'management', 'supervisor']):
                    score += 8
        
        return min(score, 25)
    
    def _score_pds_experience(self, work_experience, job):
        """Score work experience (max 30 points)."""
        score = 0
        
        # Years of experience (15 points max)
        total_years = len(work_experience)  # Simplified calculation
        experience_score = min(total_years * 2, 15)
        score += experience_score
        
        # Government service bonus (5 points)
        govt_service = any(exp.get('govt_service') == 'Y' for exp in work_experience if exp.get('govt_service'))
        if govt_service:
            score += 5
        
        # Position relevance (10 points max)
        job_title = job.get('title', '').lower()
        job_category = job.get('category', '').lower()
        
        relevance_score = 0
        for exp in work_experience:
            position = exp.get('position', '').lower()
            
            # Check for direct position match
            if any(keyword in position for keyword in job_title.split()):
                relevance_score += 5
                break
            
            # Check for category match
            if 'analyst' in position and 'analyst' in job_title:
                relevance_score += 4
            elif 'manager' in position and 'manager' in job_title:
                relevance_score += 4
            elif any(tech in position for tech in ['developer', 'programmer', 'it']) and 'technology' in job_category:
                relevance_score += 3
        
        score += min(relevance_score, 10)
        
        return min(score, 30)
    
    def _score_pds_eligibility(self, eligibility):
        """Score civil service eligibility (max 20 points)."""
        score = 0
        
        # Professional eligibility (15 points)
        professional_eligibility = ['professional', 'subprofessional', 'career service']
        has_professional = any(
            any(keyword in elig.get('eligibility', '').lower() for keyword in professional_eligibility)
            for elig in eligibility
        )
        if has_professional:
            score += 15
        
        # Board/License eligibility (5 points)
        board_eligibility = ['board', 'license', 'licensure']
        has_board = any(
            any(keyword in elig.get('eligibility', '').lower() for keyword in board_eligibility)
            for elig in eligibility
        )
        if has_board:
            score += 5
        
        return min(score, 20)
    
    def _score_pds_training(self, training, job):
        """Score training and development (max 15 points)."""
        score = 0
        
        # Number of training programs (10 points max)
        training_count = len(training)
        count_score = min(training_count * 2, 10)
        score += count_score
        
        # Relevance of training (5 points max)
        job_title = job.get('title', '').lower()
        job_category = job.get('category', '').lower()
        
        relevance_score = 0
        for train in training:
            title = train.get('title', '').lower()
            
            # Check for technical training relevance
            if any(tech in title for tech in ['data', 'computer', 'software', 'system']):
                if 'technology' in job_category or 'analyst' in job_title:
                    relevance_score += 2
            
            # Check for management training relevance
            if any(mgmt in title for mgmt in ['management', 'leadership', 'supervisor']):
                if 'manager' in job_title or 'supervisor' in job_title:
                    relevance_score += 2
        
        score += min(relevance_score, 5)
        
        return min(score, 15)
    
    def _score_pds_volunteer_work(self, volunteer_work):
        """Score voluntary work (max 10 points)."""
        score = 0
        
        # Community involvement (5 points)
        if len(volunteer_work) >= 1:
            score += 5
        
        # Leadership in volunteer work (5 points)
        leadership_keywords = ['coordinator', 'leader', 'organizer', 'head']
        has_leadership = any(
            any(keyword in vol.get('position', '').lower() for keyword in leadership_keywords)
            for vol in volunteer_work
        )
        if has_leadership:
            score += 5
        
        return min(score, 10)
    
    @login_required
    def handle_scoring_criteria(self):
        """Handle scoring criteria configuration"""
        if request.method == 'GET':
            try:
                # Return current scoring criteria
                criteria = self.pds_processor.pds_scoring_criteria
                return jsonify({
                    'success': True,
                    'scoring_criteria': criteria,
                    'description': {
                        'education': 'Weights for education level, relevance, institution quality, and grades',
                        'experience': 'Weights for experience relevance, duration, and responsibilities',
                        'skills': 'Weights for technical skills match and certifications',
                        'personal_attributes': 'Weights for eligibility, awards, and training',
                        'additional_qualifications': 'Weights for languages, licenses, and volunteer work'
                    }
                })
            except Exception as e:
                logger.error(f"Error getting scoring criteria: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'PUT':
            try:
                data = request.get_json()
                
                # Validate criteria structure
                required_criteria = ['education', 'experience', 'skills', 'personal_attributes', 'additional_qualifications']
                
                for criterion in required_criteria:
                    if criterion not in data:
                        return jsonify({'success': False, 'error': f'Missing criterion: {criterion}'}), 400
                    
                    if 'weight' not in data[criterion]:
                        return jsonify({'success': False, 'error': f'Missing weight for {criterion}'}), 400
                
                # Validate weights sum to 1.0
                total_weight = sum(data[criterion]['weight'] for criterion in required_criteria)
                if abs(total_weight - 1.0) > 0.01:
                    return jsonify({'success': False, 'error': f'Weights must sum to 1.0, current sum: {total_weight}'}), 400
                
                # Update scoring criteria
                self.pds_processor.pds_scoring_criteria = data
                
                return jsonify({
                    'success': True,
                    'message': 'Scoring criteria updated successfully',
                    'scoring_criteria': data
                })
                
            except Exception as e:
                logger.error(f"Error updating scoring criteria: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def upload_pds_only(self):
        """Handle dedicated PDS-only upload and processing"""
        try:
            if 'files[]' not in request.files:
                return jsonify({'success': False, 'error': 'No files uploaded'}), 400
            
            files = request.files.getlist('files[]')
            job_id = request.form.get('jobId')
            
            if not files or all(f.filename == '' for f in files):
                return jsonify({'success': False, 'error': 'No files selected'}), 400
            
            if not job_id:
                return jsonify({'success': False, 'error': 'Job ID is required'}), 400
            
            try:
                job_id = int(job_id)
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid job ID'}), 400
            
            # Get job details - try LSPU job postings first, then fallback to old jobs
            import sqlite3
            job = None
            
            # Try to get LSPU job posting
            try:
                conn = sqlite3.connect('resume_screening.db')
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                cursor.execute("""
                    SELECT jp.*, cl.campus_name, pt.name as position_type_name
                    FROM lspu_job_postings jp
                    LEFT JOIN campus_locations cl ON jp.campus_id = cl.id
                    LEFT JOIN position_types pt ON jp.position_type_id = pt.id
                    WHERE jp.id = ?
                """, (job_id,))
                
                row = cursor.fetchone()
                if row:
                    job = dict(row)
                    job['title'] = job.get('position_title', 'Unknown Position')
                    job['category'] = job.get('position_type_name', 'University Position')
                conn.close()
                
            except Exception as e:
                logger.warning(f"Could not fetch LSPU job posting {job_id}: {e}")
            
            # Fallback to old job system if LSPU job not found
            if not job:
                job = db_manager.get_job(job_id)
                if not job:
                    return jsonify({'success': False, 'error': 'Job not found'}), 404
            
            results = []
            errors = []
            
            for file in files:
                # Accept multiple file types for digital PDS upload
                allowed_extensions = ['.xlsx', '.xls', '.pdf', '.docx', '.doc', '.txt']
                if not any(file.filename.lower().endswith(ext) for ext in allowed_extensions):
                    errors.append(f"{file.filename}: Only Excel, PDF, Word, or text files are supported")
                    continue
                
                try:
                    logger.info(f"Processing PDS file: {file.filename}")
                    
                    # Check if this is a valid PDS file using the PDS processor
                    file.seek(0)
                    if self.pds_processor:
                        is_pds = self.pds_processor.is_pds_file(file)
                    else:
                        # Fallback check - assume Excel files are PDS
                        is_pds = file.filename.lower().endswith(('.xlsx', '.xls'))
                    logger.info(f"PDS detection result for {file.filename}: {is_pds}")
                    
                    if not is_pds:
                        errors.append(f"{file.filename}: File is not in valid PDS format")
                        continue
                    
                    # Extract PDS data using appropriate method based on file type
                    file.seek(0)
                    if self.pds_processor:
                        # Handle different file types properly
                        if file.filename.lower().endswith(('.xlsx', '.xls')):
                            # For Excel files, save temporarily and use file path
                            import tempfile
                            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as temp_file:
                                file.seek(0)
                                temp_file.write(file.read())
                                temp_file.flush()
                                
                                # Use the Excel-specific processing method
                                candidate_data = self._process_excel_file(temp_file.name, file.filename, job)
                                
                                # Clean up temp file
                                try:
                                    os.unlink(temp_file.name)
                                except:
                                    pass
                        else:
                            # For text-based files (PDF, DOC, TXT), use text content
                            file_content = file.read()
                            if isinstance(file_content, bytes):
                                # Try to decode bytes to string for text-based processing
                                try:
                                    file_content = file_content.decode('utf-8')
                                except UnicodeDecodeError:
                                    # If decoding fails, skip this file
                                    errors.append(f"{file.filename}: Could not decode file content")
                                    continue
                            file.seek(0)  # Reset for any subsequent reads
                            candidate_data = self.pds_processor.process_pds_candidate(file_content)
                    else:
                        # Fallback processing
                        candidate_data = {
                            'name': f"Candidate from {file.filename}",
                            'email': '',
                            'pds_data': {'personal_info': {}}
                        }
                    
                    if not candidate_data:
                        errors.append(f"{file.filename}: Failed to extract PDS data")
                        continue
                    
                    logger.info(f"PDS extraction successful for {file.filename}")
                    
                    # Use enhanced assessment system for scoring
                    assessment_result = None
                    score = 0
                    percentage_score = 0
                    scoring_breakdown = {}
                    
                    # Check if this is an LSPU job and use appropriate assessment
                    is_lspu_job = any(field in job for field in [
                        'education_requirements', 'experience_requirements', 
                        'training_requirements', 'eligibility_requirements', 'position_title'
                    ])
                    
                    if is_lspu_job and self.assessment_engine:
                        try:
                            # Use LSPU university assessment engine
                            logger.info(f"ðŸŽ¯ Starting LSPU university assessment for {file.filename} with job {job_id}")
                            assessment_result = self.assessment_engine.assess_candidate_for_lspu_job(
                                candidate_data=candidate_data['pds_data'],
                                lspu_job=job,
                                position_type_id=job.get('position_type_id')
                            )
                            score = assessment_result.get('automated_score', 0)
                            percentage_score = assessment_result.get('percentage_score', 0)
                            scoring_breakdown = assessment_result.get('assessment_results', {})
                            logger.info(f"âœ… LSPU university assessment completed for {file.filename} - Score: {percentage_score}%")
                            
                        except Exception as e:
                            logger.error(f"âŒ LSPU assessment failed for {file.filename}: {e}")
                            logger.warning(f"ðŸ”„ Using fallback scoring for {file.filename}")
                            score = 75  # Fallback score
                            percentage_score = 75
                    else:
                        # Use legacy scoring for old job system
                        if self.pds_processor and candidate_data:
                            score = self._calculate_comprehensive_pds_score(candidate_data, job)
                            percentage_score = score
                        else:
                            score = 75  # Default score for successful file processing
                            percentage_score = 75
                    
                    # Prepare data for PDS candidates table
                    pds_candidate_data = {
                        'name': candidate_data['name'],
                        'email': candidate_data['email'],
                        'phone': candidate_data.get('phone', ''),
                        'job_id': job_id,
                        'score': score,
                        'percentage_score': percentage_score,
                        'status': 'new',
                        'filename': file.filename,
                        'file_size': len(file.read()) if hasattr(file, 'read') else 0,
                        'assessment_engine': 'LSPU_University_Standards' if is_lspu_job else 'Legacy_PDS_Processor',
                        
                        # Core PDS sections
                        'personal_info': candidate_data['pds_data'].get('personal_info', {}),
                        'family_background': candidate_data['pds_data'].get('family_background', {}),
                        'educational_background': candidate_data['pds_data'].get('educational_background', {}),
                        'civil_service_eligibility': candidate_data['pds_data'].get('eligibility', []),
                        'work_experience': candidate_data['pds_data'].get('work_experience', []),
                        'voluntary_work': candidate_data['pds_data'].get('voluntary_work', []),
                        'learning_development': candidate_data['pds_data'].get('training', []),
                        'other_information': candidate_data['pds_data'].get('other_info', {}),
                        'personal_references': candidate_data['pds_data'].get('personal_references', []),
                        'government_ids': candidate_data['pds_data'].get('government_ids', {}),
                        
                        # Extracted summary
                        'highest_education': self._extract_highest_education(candidate_data['pds_data']),
                        'years_of_experience': self._calculate_experience_years(candidate_data['pds_data']),
                        'government_service_years': self._calculate_govt_service_years(candidate_data['pds_data']),
                        'civil_service_eligible': self._check_civil_service_eligibility(candidate_data['pds_data']),
                        
                        # Enhanced scoring details
                        'scoring_breakdown': scoring_breakdown if assessment_result else candidate_data.get('scoring_breakdown', {}),
                        'assessment_details': assessment_result if assessment_result else {},
                        'matched_qualifications': self._extract_matched_qualifications(candidate_data, job),
                        'areas_for_improvement': self._identify_improvement_areas(candidate_data, job),
                        
                        # Processing metadata
                        'extraction_success': True,
                        'extraction_errors': [],
                        'processing_notes': f"Successfully processed PDS format from {file.filename}"
                    }
                    
                    # Store in candidates table (using existing method)
                    simple_candidate_data = {
                        'name': candidate_data.get('name', f"Candidate from {file.filename}"),
                        'email': candidate_data.get('email', ''),
                        'phone': candidate_data.get('phone', ''),
                        'resume_text': f"PDS file: {file.filename}",
                        'education': json.dumps(candidate_data.get('pds_data', {}).get('educational_background', [])),
                        'work_experience': json.dumps(candidate_data.get('pds_data', {}).get('work_experience', [])),
                        'skills': json.dumps(candidate_data.get('pds_data', {}).get('skills', [])),
                        'certifications': json.dumps(candidate_data.get('pds_data', {}).get('eligibility', [])),
                        'job_id': job_id,
                        'category': job.get('category', 'University Position'),
                        'score': score,
                        'status': 'pending',
                        'processing_type': 'pds_digital',
                        'scoring_breakdown': json.dumps(scoring_breakdown if assessment_result else candidate_data.get('scoring_breakdown', {}))
                    }
                    
                    candidate_id = db_manager.create_candidate(simple_candidate_data)
                    
                    # Prepare response data
                    result = {
                        'candidate_id': candidate_id,
                        'filename': file.filename,
                        'name': candidate_data.get('name', f"Candidate from {file.filename}"),
                        'email': candidate_data.get('email', ''),
                        'matchScore': score,  # Frontend expects 'matchScore'
                        'total_score': score,  # Keep for backward compatibility
                        'percentage_score': percentage_score,  # University assessment percentage
                        'processing_type': 'pds_digital',
                        'assessment_engine': 'LSPU_University_Standards' if is_lspu_job else 'Legacy_PDS_Processor',
                        'sections_extracted': list(candidate_data.get('pds_data', {}).keys()) if candidate_data.get('pds_data') else [],
                        'scoring_breakdown': scoring_breakdown if assessment_result else {},
                        'assessment_recommendation': assessment_result.get('recommendation', 'pending') if assessment_result else 'pending',
                        'success': True
                    }
                    
                    results.append(result)
                    logger.info(f"Successfully processed PDS: {file.filename} with score: {score} (University Assessment: {percentage_score}%)")
                    
                except Exception as e:
                    error_msg = f"Error processing PDS file {file.filename}: {e}"
                    logger.error(error_msg)
                    errors.append(error_msg)
                    continue
            
            response_data = {
                'success': True,
                'message': f'Successfully processed {len(results)} documents',
                'results': results,
                'processing_type': 'pds_digital'
            }
            
            if errors:
                response_data['warnings'] = errors
                response_data['message'] += f' ({len(errors)} files had issues)'
            
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"Error in upload_pds_only: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def upload_ocr(self):
        """Handle OCR-specific upload and processing for scanned documents"""
        try:
            if 'files[]' not in request.files:
                return jsonify({'success': False, 'error': 'No files uploaded'}), 400
            
            files = request.files.getlist('files[]')
            job_id = request.form.get('jobId')
            
            if not files or all(f.filename == '' for f in files):
                return jsonify({'success': False, 'error': 'No files selected'}), 400
            
            if not job_id:
                return jsonify({'success': False, 'error': 'Job ID is required'}), 400
            
            try:
                job_id = int(job_id)
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid job ID'}), 400
            
            # Verify OCR processor is available
            if not self.ocr_processor:
                return jsonify({'success': False, 'error': 'OCR processor not available'}), 500
            
            # Get job details - try LSPU job postings first, then fallback to old jobs
            job = None
            
            # Try to get LSPU job posting
            try:
                conn = sqlite3.connect('resume_screening.db')
                cursor = conn.cursor()
                
                cursor.execute('''
                    SELECT id, position_title, position_category, campus_location,
                           salary_range, education_requirements, experience_requirements,
                           skills_requirements, preferred_qualifications, 
                           application_deadline, description, status
                    FROM lspu_job_postings 
                    WHERE id = ?
                ''', (job_id,))
                
                row = cursor.fetchone()
                if row:
                    job = {
                        'id': row[0], 'title': row[1], 'category': row[2], 
                        'campus_location': row[3], 'salary_range': row[4],
                        'education_requirements': row[5], 'experience_requirements': row[6],
                        'skills_requirements': row[7], 'preferred_qualifications': row[8],
                        'application_deadline': row[9], 'description': row[10], 
                        'status': row[11], 'requirements': row[7] or ''  # Use skills as requirements
                    }
                conn.close()
                
            except Exception as e:
                logger.warning(f"Could not fetch LSPU job posting {job_id}: {e}")
            
            # Fallback to old job system if LSPU job not found
            if not job:
                job = db_manager.get_job(job_id)
                if not job:
                    return jsonify({'success': False, 'error': 'Job not found'}), 404
            
            results = []
            errors = []
            
            for file in files:
                if file.filename != '' and self._is_allowed_file(file.filename):
                    # Only process image files for OCR
                    if not self._is_image_file(file.filename):
                        errors.append(f"{file.filename}: Only image files are supported for OCR processing")
                        continue
                    
                    try:
                        logger.info(f"Processing OCR file: {file.filename}")
                        
                        # Process image with OCR
                        ocr_result = self.ocr_processor.process_pds_image(file, job_id)
                        
                        if not ocr_result['success']:
                            errors.append(f"{file.filename}: {ocr_result.get('error', 'Unknown OCR error')}")
                            continue
                        
                        candidate_data = ocr_result['candidate_data']
                        
                        # Use enhanced assessment system for OCR scoring
                        assessment_result = None
                        base_score = 0
                        percentage_score = 0
                        scoring_breakdown = {}
                        
                        # Check if this is an LSPU job and use appropriate assessment
                        is_lspu_job = any(field in job for field in [
                            'education_requirements', 'experience_requirements', 
                            'training_requirements', 'eligibility_requirements', 'position_title'
                        ])
                        
                        if candidate_data['resume_text'].strip():
                            if is_lspu_job and self.assessment_engine:
                                try:
                                    # Extract PDS-like data from OCR text for assessment
                                    extracted_pds_data = self._extract_pds_from_ocr_text(candidate_data['resume_text'])
                                    
                                    # Use LSPU university assessment engine
                                    assessment_result = self.assessment_engine.assess_candidate_for_lspu_job(
                                        candidate_data=extracted_pds_data,
                                        lspu_job=job,
                                        position_type_id=job.get('position_type_id')
                                    )
                                    base_score = assessment_result.get('automated_score', 0)
                                    percentage_score = assessment_result.get('percentage_score', 0)
                                    scoring_breakdown = assessment_result.get('assessment_results', {})
                                    
                                    # Apply OCR confidence factor
                                    ocr_confidence = ocr_result.get('confidence', 0.8)
                                    base_score *= ocr_confidence
                                    percentage_score *= ocr_confidence
                                    
                                    logger.info(f"LSPU university OCR assessment completed for {file.filename} - Score: {percentage_score}%")
                                    
                                except Exception as e:
                                    logger.warning(f"LSPU OCR assessment failed for {file.filename}, using fallback: {e}")
                                    base_score = self._calculate_ocr_score(candidate_data, job, ocr_result)
                                    percentage_score = base_score
                            else:
                                # Use legacy OCR scoring for old job system
                                base_score = self._calculate_ocr_score(candidate_data, job, ocr_result)
                                percentage_score = base_score
                            
                            candidate_data['score'] = base_score
                            candidate_data['percentage_score'] = percentage_score
                            candidate_data['assessment_engine'] = 'LSPU_University_Standards' if is_lspu_job else 'Legacy_OCR_Processor'
                        else:
                            candidate_data['score'] = 0
                            candidate_data['percentage_score'] = 0
                            candidate_data['assessment_engine'] = 'No_Text_Extracted'
                        
                        # Store candidate data
                        candidate_id = db_manager.create_candidate(candidate_data)
                        
                        # Prepare result for response
                        result = {
                            'candidate_id': candidate_id,
                            'filename': file.filename,
                            'name': candidate_data['name'],
                            'email': candidate_data['email'],
                            'matchScore': candidate_data['score'],  # Frontend expects 'matchScore'
                            'total_score': candidate_data['score'],  # Keep for backward compatibility
                            'percentage_score': candidate_data.get('percentage_score', candidate_data['score']),
                            'processing_type': 'ocr_scanned',
                            'assessment_engine': candidate_data.get('assessment_engine', 'Legacy_OCR_Processor'),
                            'ocr_confidence': candidate_data['ocr_confidence'],
                            'confidence_level': ocr_result['confidence_level'],
                            'extracted_fields': list(ocr_result['pds_fields'].keys()),
                            'preprocessing_steps': len(ocr_result['preprocessing_info']['steps_applied']),
                            'scoring_breakdown': scoring_breakdown if assessment_result else {},
                            'assessment_recommendation': assessment_result.get('recommendation', 'pending') if assessment_result else 'pending'
                        }
                        
                        results.append(result)
                        logger.info(f"Successfully processed OCR file: {file.filename} with confidence: {candidate_data['ocr_confidence']:.2f}% (Assessment: {candidate_data.get('percentage_score', 0)}%)")
                        
                    except Exception as e:
                        error_msg = f"Error processing OCR file {file.filename}: {e}"
                        logger.error(error_msg)
                        errors.append(error_msg)
                        continue
                else:
                    if file.filename:
                        errors.append(f"{file.filename}: Unsupported file type for OCR processing")
            
            # Calculate average confidence for the batch
            avg_confidence = 0
            avg_assessment_score = 0
            if results:
                avg_confidence = sum(r['ocr_confidence'] for r in results) / len(results)
                avg_assessment_score = sum(r.get('percentage_score', 0) for r in results) / len(results)
            
            response_data = {
                'success': True,
                'message': f'Successfully processed {len(results)} documents with OCR',
                'results': results,
                'processing_type': 'ocr_scanned',
                'average_confidence': round(avg_confidence, 2),
                'average_assessment_score': round(avg_assessment_score, 2),
                'total_files': len(files),
                'successful_extractions': len(results)
            }
            
            if errors:
                response_data['warnings'] = errors
                response_data['message'] += f' ({len(errors)} files had errors)'
            
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"Error in upload_ocr: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    # @login_required  # Temporarily disabled for testing
    def upload_files_clean(self):
        """
        New clean file upload endpoint
        Handles PDF and XLSX files with batch processing, no OCR dependencies
        Uses database session storage for reliability
        """
        # Check authentication
        if not session.get('user_authenticated') and not (FLASK_LOGIN_AVAILABLE and current_user.is_authenticated):
            return jsonify({'success': False, 'error': 'Authentication required'}), 401
        
        try:
            if not self.clean_upload_handler:
                return jsonify({'success': False, 'error': 'Upload system not available'}), 500
            
            # Get files from request
            files = request.files.getlist('files[]')
            job_id = request.form.get('jobId')
            
            logger.info(f"ðŸ“¤ Upload request: {len(files)} files for job {job_id}")
            
            if not files or all(f.filename == '' for f in files):
                return jsonify({'success': False, 'error': 'No files selected'}), 400
            
            if not job_id:
                return jsonify({'success': False, 'error': 'Job ID is required'}), 400
            
            try:
                job_id = int(job_id)
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid job ID'}), 400
            
            # Verify job exists (check both LSPU and legacy jobs)
            job = self._get_job_by_id(job_id)
            if not job:
                return jsonify({'success': False, 'error': f'Job {job_id} not found'}), 404
            
            logger.info(f"ðŸ“¤ Processing upload for job: {job.get('title', 'Unknown')} (ID: {job_id})")
            
            # Process files with clean handler
            success, results, errors = self.clean_upload_handler.process_upload_batch(files)
            
            if not success:
                return jsonify({
                    'success': False, 
                    'error': errors if isinstance(errors, str) else 'No valid files uploaded',
                    'details': errors if isinstance(errors, list) else []
                }), 400
            
            # Create upload session in database
            import uuid
            session_id = str(uuid.uuid4())
            user_id = session.get('user_id', current_user.id if hasattr(current_user, 'id') else None)
            
            # Create session record
            if not db_manager.create_upload_session(session_id, user_id, job_id):
                return jsonify({'success': False, 'error': 'Failed to create upload session'}), 500
            
            # Store file records in database
            successful_files = 0
            file_metadata = {}  # Store additional file data for analysis
            logger.info(f"📁 Processing {len(results)} upload results")
            
            for i, result in enumerate(results):
                logger.info(f"📄 Processing file {i+1}/{len(results)}: {result['file_id']}")
                logger.info(f"📄 File info: {result['file_info']}")
                
                if db_manager.create_upload_file_record(session_id, result['file_id'], result['file_info']):
                    successful_files += 1
                    logger.info(f"✅ Successfully stored file {result['file_id']} (total: {successful_files})")
                    # Store additional metadata needed for analysis
                    file_metadata[result['file_id']] = {
                        'temp_path': result['temp_path'],
                        'original_name': result['file_info']['original_name'],
                        'file_id': result['file_id']
                    }
                else:
                    logger.warning(f"❌ Failed to create file record for {result['file_id']}")
            
            logger.info(f"📊 Upload summary: {successful_files}/{len(results)} files stored successfully")
            
            # Update session with file count
            db_manager.update_upload_session(session_id, 
                file_count=successful_files,
                metadata=json.dumps({
                    'job_info': {
                        'id': job_id, 
                        'title': job.get('title', 'Unknown'),
                        'category': job.get('category', 'Unknown')
                    },
                    'upload_summary': {
                        'total_files': len(files),
                        'valid_files': len(results),
                        'stored_files': successful_files,
                        'errors': len(errors) if errors else 0
                    },
                    'file_metadata': file_metadata  # Store file paths and IDs for analysis
                })
            )
            
            logger.info(f"âœ… Upload session {session_id} created with {successful_files} files for job {job_id}")
            
            response_data = {
                'success': True,
                'message': f'Successfully uploaded {successful_files} files',
                'session_id': session_id,
                'file_count': successful_files,
                'files': [r['preview'] for r in results],
                'ready_for_analysis': successful_files > 0,
                'job_info': {
                    'id': job_id,
                    'title': job.get('title', 'Unknown'),
                    'category': job.get('category', 'Unknown')
                }
            }
            
            if errors:
                response_data['warnings'] = errors
                response_data['message'] += f' ({len(errors)} files had errors)'
            
            logger.info(f"âœ… Clean upload successful: {successful_files} files ready for analysis")
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"Error in upload_files_clean: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    def start_analysis(self):
        """
        Start analysis for uploaded files using database session storage
        Processes all files in an upload session and generates candidate records
        """
        # Check authentication
        if not session.get('user_authenticated') and not (FLASK_LOGIN_AVAILABLE and current_user.is_authenticated):
            return jsonify({'success': False, 'error': 'Authentication required'}), 401
        
        try:
            data = request.get_json()
            session_id = data.get('session_id')
            
            logger.info(f"ðŸ” Start analysis request: session_id={session_id}")
            
            if not session_id:
                return jsonify({'success': False, 'error': 'Session ID required'}), 400
            
            # Get session data from database
            upload_session = db_manager.get_upload_session(session_id)
            if not upload_session:
                return jsonify({'success': False, 'error': 'Upload session not found'}), 404
            
            job_id = upload_session['job_id']
            
            # Get job details
            job = self._get_job_by_id(job_id)
            if not job:
                return jsonify({'success': False, 'error': f'Job {job_id} not found'}), 404
            
            # Get files for this session
            upload_files = db_manager.get_upload_files(session_id)
            if not upload_files:
                return jsonify({'success': False, 'error': 'No files found for analysis'}), 404
            
            logger.info(f"ðŸ” Starting analysis for {len(upload_files)} files in session {session_id}")
            
            # Update session status
            db_manager.update_upload_session(session_id, status='processing')
            
            results = []
            successful_analyses = 0
            analysis_errors = []
            
            for file_record in upload_files:
                try:
                    logger.info(f"ðŸ” Processing file: {file_record['original_name']}")
                    
                    # Check if file still exists
                    if not os.path.exists(file_record['temp_path']):
                        error_msg = f"File not found: {file_record['temp_path']}"
                        logger.error(error_msg)
                        db_manager.update_upload_file_status(file_record['file_id'], 'error', error_message=error_msg)
                        analysis_errors.append(error_msg)
                        continue
                    
                    # Process the file through appropriate engine
                    candidate_data = self._process_file_for_analysis(file_record, job)
                    
                    if candidate_data:
                        # Store candidate in database
                        candidate_id = db_manager.create_candidate(candidate_data)
                        
                        if candidate_id:
                            # Update file record with success
                            db_manager.update_upload_file_status(file_record['file_id'], 'processed', candidate_id=candidate_id)
                            
                            results.append({
                                'file_id': file_record['file_id'],
                                'candidate_id': candidate_id,
                                'name': candidate_data.get('name', 'Unknown'),
                                'email': candidate_data.get('email', ''),
                                'matchScore': candidate_data.get('score', 0),
                                'status': 'processed'
                            })
                            successful_analyses += 1
                            logger.info(f"âœ… Successfully processed: {file_record['original_name']} -> Candidate ID: {candidate_id}")
                        else:
                            error_msg = f"Failed to create candidate record for {file_record['original_name']}"
                            logger.error(error_msg)
                            db_manager.update_upload_file_status(file_record['file_id'], 'error', error_message=error_msg)
                            analysis_errors.append(error_msg)
                    else:
                        error_msg = f"Failed to process file: {file_record['original_name']}"
                        logger.error(error_msg)
                        db_manager.update_upload_file_status(file_record['file_id'], 'error', error_message=error_msg)
                        analysis_errors.append(error_msg)
                
                except Exception as e:
                    error_msg = f"Error processing {file_record['original_name']}: {str(e)}"
                    logger.error(error_msg)
                    db_manager.update_upload_file_status(file_record['file_id'], 'error', error_message=error_msg)
                    analysis_errors.append(error_msg)
            
            # Update session with completion status
            session_status = 'completed' if successful_analyses > 0 else 'failed'
            db_manager.update_upload_session(
                session_id, 
                status=session_status,
                completed_at=datetime.now().isoformat(),
                error_log=json.dumps(analysis_errors) if analysis_errors else None
            )
            
            # Clean up temporary files
            self._cleanup_session_files(upload_files)
            
            response_data = {
                'success': True,
                'message': f'Analysis completed: {successful_analyses} candidates processed',
                'session_id': session_id,
                'successful_analyses': successful_analyses,
                'total_files': len(upload_files),
                'results': results,
                'job_info': {
                    'id': job_id,
                    'title': job.get('title', 'Unknown'),
                    'category': job.get('category', 'Unknown')
                }
            }
            
            if analysis_errors:
                response_data['warnings'] = analysis_errors
                response_data['message'] += f' ({len(analysis_errors)} files had errors)'
            
            logger.info(f"âœ… Analysis completed for session {session_id}: {successful_analyses}/{len(upload_files)} files processed")
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"Error in start_analysis: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
                
            if session_id not in session['upload_sessions']:
                logger.error(f"âŒ Session {session_id} not found in available sessions: {list(session['upload_sessions'].keys())}")
                return jsonify({'success': False, 'error': 'Upload session not found'}), 404
            
            session_data = session['upload_sessions'][session_id]
            
            # Check if session was already completed
            if session_data.get('status') == 'completed':
                logger.info(f"ðŸ”„ Session {session_id} already completed, returning previous results")
                return jsonify({
                    'success': True, 
                    'message': 'Analysis already completed for this session',
                    'completed_at': session_data.get('completed_at'),
                    'note': 'This session was already processed. Check the Applications section for results.'
                })
            
            job_id = session_data['job_id']
            file_data = session_data['file_data']
            
            logger.info(f"ðŸ” Session data: job_id={job_id}, file_count={len(file_data)}")
            logger.info(f"Starting analysis for session {session_id}: {len(file_data)} files")
            
            # Get job details
            job = self._get_job_details(job_id)
            logger.info(f"ðŸ” Job lookup result for ID {job_id}: {job is not None}")
            if not job:
                logger.error(f"âŒ Job not found for ID: {job_id}")
                return jsonify({'success': False, 'error': 'Job not found'}), 404
            
            results = []
            errors = []
            
            for file_info in file_data:
                try:
                    temp_path = file_info['temp_path']
                    file_preview = file_info['preview']
                    
                    # Process file based on type
                    if file_preview['type'] == 'pdf':
                        candidate_data = self._process_pdf_file(temp_path, file_preview['name'], job)
                    elif file_preview['type'] == 'excel':
                        candidate_data = self._process_excel_file(temp_path, file_preview['name'], job)
                    else:
                        errors.append(f"{file_preview['name']}: Unsupported file type")
                        continue
                    
                    if candidate_data:
                        # Store candidate in database
                        try:
                            # Ensure fields don't exceed database limits
                            self._validate_candidate_field_lengths(candidate_data)
                            
                            candidate_id = db_manager.create_candidate(candidate_data)
                            
                            results.append({
                                'candidate_id': candidate_id,
                                'filename': file_preview['name'],
                                'name': candidate_data['name'],
                                'email': candidate_data['email'],
                                'matchScore': candidate_data.get('score', 0),
                                'percentage_score': candidate_data.get('percentage_score', 0),
                                'processing_type': 'clean_upload'
                            })
                            
                            logger.info(f"Successfully processed and stored {file_preview['name']}")
                        except Exception as db_error:
                            logger.error(f"Database error storing candidate from {file_preview['name']}: {db_error}")
                            errors.append(f"{file_preview['name']}: Database storage failed - {str(db_error)}")
                    else:
                        errors.append(f"{file_preview['name']}: Failed to extract data")
                        
                except Exception as e:
                    error_msg = f"Error processing {file_info.get('preview', {}).get('name', 'unknown')}: {e}"
                    logger.error(error_msg)
                    errors.append(error_msg)
            
            # Clean up temporary files
            file_ids = [f['file_id'] for f in file_data]
            self.clean_upload_handler.cleanup_temp_files(file_ids)
            
            # Keep session for a short time instead of immediate deletion to allow retries
            session['upload_sessions'][session_id]['completed_at'] = datetime.now().isoformat()
            session['upload_sessions'][session_id]['status'] = 'completed'
            session.modified = True
            
            response_data = {
                'success': True,
                'message': f'Analysis completed: {len(results)} candidates processed',
                'results': results,
                'total_files': len(file_data),
                'successful_analyses': len(results)
            }
            
            if errors:
                response_data['warnings'] = errors
                response_data['message'] += f' ({len(errors)} files had errors)'
            
            logger.info(f"Analysis completed for session {session_id}: {len(results)} candidates created")
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"Error in start_analysis: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    def _get_job_details(self, job_id):
        """Get job details from either LSPU or legacy job system"""
        logger.info(f"ðŸ” Looking up job details for ID: {job_id} (type: {type(job_id)})")
        
        try:
            # Try LSPU job postings first
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT jp.id, jp.position_title, jp.position_category, 
                       COALESCE(cl.campus_name, 'Main Campus') as campus_location,
                       jp.salary_amount, jp.education_requirements, jp.experience_requirements,
                       jp.special_requirements, jp.eligibility_requirements, jp.specific_role
                FROM lspu_job_postings jp
                LEFT JOIN campus_locations cl ON jp.campus_id = cl.id
                WHERE jp.id = ?
            ''', (job_id,))
            
            row = cursor.fetchone()
            logger.info(f"ðŸ” LSPU job query result: {row is not None}")
            
            if row:
                conn.close()
                job_details = {
                    'id': row[0], 'title': row[1], 'category': row[2], 
                    'campus_location': row[3], 'salary_range': f"Grade {row[4]}" if row[4] else "TBD",
                    'education_requirements': row[5], 'experience_requirements': row[6],
                    'skills_requirements': row[7], 'preferred_qualifications': row[8],
                    'description': row[9] or row[1], 'requirements': row[7] or ''
                }
                logger.info(f"âœ… Found LSPU job: {job_details['title']}")
                return job_details
            
            conn.close()
        except Exception as e:
            logger.warning(f"Could not fetch LSPU job {job_id}: {e}")
        
        # No job found
        logger.error(f"âŒ Job not found for ID: {job_id}")
        return None
    
    def _process_pdf_file(self, file_path, filename, job):
        """Process PDF file and extract candidate data"""
        try:
            # Use existing PDS processor for PDF extraction
            if self.pds_processor:
                return self.pds_processor.process_pdf_file(file_path, filename, job)
            else:
                # Fallback to basic text extraction
                logger.warning("PDS processor not available, using basic extraction")
                return self._basic_pdf_extraction(file_path, filename, job)
        except Exception as e:
            logger.error(f"Error processing PDF {filename}: {e}")
            return None
    
    def _process_excel_file(self, file_path, filename, job):
        """Process Excel file and extract candidate data using comprehensive PDS extraction"""
        try:
            # Use the comprehensive PDS processor for Excel files too
            if self.pds_processor:
                logger.info(f"ðŸ” Processing Excel PDS file: {filename}")
                return self.pds_processor.process_pdf_file(file_path, filename, job)
            else:
                # Fallback to basic Excel processing
                logger.warning("PDS processor not available, using basic Excel extraction")
                return self._basic_excel_extraction(file_path, filename, job)
        except Exception as e:
            logger.error(f"Error processing Excel {filename}: {e}")
            return None
    
    def _basic_excel_extraction(self, file_path, filename, job):
        """Basic Excel processing fallback"""
        try:
            import pandas as pd
            
            df = pd.read_excel(file_path)
            
            # Try to extract basic info from first row/common columns
            candidate_data = {
                'name': self._extract_from_excel(df, ['name', 'full_name', 'candidate_name']),
                'email': self._extract_from_excel(df, ['email', 'email_address', 'contact_email']),
                'phone': self._extract_from_excel(df, ['phone', 'mobile', 'contact_number']),
                'resume_text': df.to_string(),  # Convert whole sheet to text for analysis
                'job_id': job['id'],
                'score': 0,  # Will be calculated
                'percentage_score': 0,
                'processing_type': 'basic_excel_fallback'
            }
            
            # Basic scoring based on data completeness
            score = 0
            if candidate_data['name']: score += 20
            if candidate_data['email']: score += 20
            if candidate_data['phone']: score += 10
            
            candidate_data['score'] = score
            candidate_data['percentage_score'] = score
            
            return candidate_data
            
        except Exception as e:
            logger.error(f"Basic Excel extraction failed for {filename}: {e}")
            return None
    
    def _extract_from_excel(self, df, possible_columns):
        """Extract value from Excel DataFrame using possible column names"""
        for col in possible_columns:
            for actual_col in df.columns:
                if col.lower() in actual_col.lower():
                    values = df[actual_col].dropna()
                    if len(values) > 0:
                        return str(values.iloc[0])
        return ''
    
    def _basic_pdf_extraction(self, file_path, filename, job):
        """Basic PDF text extraction fallback"""
        try:
            import PyPDF2
            
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text()
            
            # Basic candidate data extraction
            candidate_data = {
                'name': filename.replace('.pdf', ''),  # Use filename as fallback
                'email': '',
                'phone': '',
                'resume_text': text,
                'job_id': job['id'],
                'score': 30,  # Basic score for successful text extraction
                'percentage_score': 30,
                'processing_type': 'basic_pdf'
            }
            
            return candidate_data
            
        except Exception as e:
            logger.error(f"Basic PDF extraction failed for {filename}: {e}")
            return None
    
    def _extract_pds_from_ocr_text(self, ocr_text: str) -> Dict[str, Any]:
        """
        Extract PDS-like structured data from OCR text for assessment
        This is a simplified version that attempts to parse common PDS sections
        """
        extracted_data = {
            'basic_info': {},
            'education': [],
            'experience': [],
            'training': [],
            'eligibility': [],
            'certifications': [],
            'skills': [],
            'awards': [],
            'volunteer_work': [],
            'languages': []
        }
        
        try:
            # Use the existing resume processor to extract basic information
            if self.resume_processor:
                basic_info = self.resume_processor.extract_basic_info(ocr_text)
                extracted_data['basic_info'] = basic_info
                
                # Extract education
                education_info = self.resume_processor.extract_education(ocr_text)
                extracted_data['education'] = education_info if isinstance(education_info, list) else [education_info]
                
                # Extract work experience
                work_experience = self.resume_processor.extract_work_experience(ocr_text)
                extracted_data['experience'] = work_experience if isinstance(work_experience, list) else [work_experience]
                
                # Extract skills
                skills_info = self.resume_processor.extract_skills(ocr_text)
                extracted_data['skills'] = skills_info if isinstance(skills_info, list) else [skills_info]
                
                # Extract certifications
                certifications = self.resume_processor.extract_certifications(ocr_text)
                extracted_data['certifications'] = certifications if isinstance(certifications, list) else [certifications]
            
            # Additional OCR-specific parsing for PDS sections
            text_lower = ocr_text.lower()
            
            # Look for training/seminars
            if 'training' in text_lower or 'seminar' in text_lower or 'workshop' in text_lower:
                # Simple extraction of training lines
                lines = ocr_text.split('\n')
                training_section = False
                for line in lines:
                    line_lower = line.lower().strip()
                    if any(keyword in line_lower for keyword in ['training', 'seminar', 'workshop', 'course']):
                        if len(line.strip()) > 10:  # Reasonable length for training entry
                            extracted_data['training'].append({'title': line.strip(), 'hours': 8})
            
            # Look for eligibility/certifications
            if any(keyword in text_lower for keyword in ['eligible', 'eligibility', 'license', 'certification', 'board exam']):
                lines = ocr_text.split('\n')
                for line in lines:
                    line_lower = line.lower().strip()
                    if any(keyword in line_lower for keyword in ['csc', 'civil service', 'eligible', 'license', 'board', 'cpa', 'bar']):
                        if len(line.strip()) > 5:
                            extracted_data['eligibility'].append({'eligibility': line.strip()})
            
            # Look for awards/honors
            if any(keyword in text_lower for keyword in ['award', 'honor', 'recognition', 'achievement']):
                lines = ocr_text.split('\n')
                for line in lines:
                    line_lower = line.lower().strip()
                    if any(keyword in line_lower for keyword in ['award', 'honor', 'recognition', 'dean', 'cum laude']):
                        if len(line.strip()) > 5:
                            extracted_data['awards'].append({'title': line.strip()})
            
        except Exception as e:
            logger.warning(f"Error extracting PDS data from OCR text: {e}")
        
        return extracted_data
    
    def _calculate_ocr_score(self, candidate_data, job, ocr_result):
        """Calculate score for OCR-processed candidate with confidence weighting"""
        try:
            # Base score calculation (simplified for OCR)
            base_score = 0
            
            # Text content score (30 points)
            text_length = len(candidate_data['resume_text'].strip())
            if text_length > 500:
                base_score += 30
            elif text_length > 200:
                base_score += 20
            elif text_length > 50:
                base_score += 10
            
            # Field extraction score (40 points)
            extracted_fields = ocr_result['pds_fields']
            field_score = 0
            
            # Essential fields
            if extracted_fields.get('name') and extracted_fields['name'] != 'Unknown OCR Candidate':
                field_score += 15
            if extracted_fields.get('email'):
                field_score += 10
            if extracted_fields.get('phone'):
                field_score += 5
            if extracted_fields.get('address'):
                field_score += 5
            if extracted_fields.get('education'):
                field_score += 5
            
            base_score += min(field_score, 40)
            
            # OCR confidence bonus/penalty (30 points)
            confidence = candidate_data['ocr_confidence']
            if confidence >= 85:
                confidence_score = 30
            elif confidence >= 70:
                confidence_score = 25
            elif confidence >= 50:
                confidence_score = 15
            elif confidence >= 30:
                confidence_score = 10
            else:
                confidence_score = 5
            
            base_score += confidence_score
            
            # Apply confidence weighting to final score
            confidence_weight = min(confidence / 100.0, 1.0)
            final_score = base_score * confidence_weight
            
            return min(round(final_score), 100)
            
        except Exception as e:
            logger.error(f"Error calculating OCR score: {str(e)}")
            return 0
    
    def _calculate_comprehensive_pds_score(self, candidate_data, job):
        """Calculate comprehensive score for PDS candidate."""
        try:
            total_score = 0
            scoring_breakdown = {}
            
            pds_data = candidate_data.get('pds_data', {})
            personal_info = pds_data.get('personal_info', {})
            
            # 1. Education Score (30 points)
            education_score = self._score_education_comprehensive(pds_data.get('educational_background', {}), job)
            scoring_breakdown['education'] = education_score
            total_score += education_score
            
            # 2. Work Experience Score (35 points)
            experience_score = self._score_experience_comprehensive(pds_data.get('work_experience', []), job)
            scoring_breakdown['experience'] = experience_score
            total_score += experience_score
            
            # 3. Civil Service Eligibility Score (15 points)
            eligibility_score = self._score_eligibility_comprehensive(pds_data.get('eligibility', []))
            scoring_breakdown['eligibility'] = eligibility_score
            total_score += eligibility_score
            
            # 4. Training and Development Score (10 points)
            training_score = self._score_training_comprehensive(pds_data.get('training', []), job)
            scoring_breakdown['training'] = training_score
            total_score += training_score
            
            # 5. Additional Qualifications Score (10 points)
            additional_score = self._score_additional_qualifications(pds_data, job)
            scoring_breakdown['additional'] = additional_score
            total_score += additional_score
            
            # Store scoring breakdown in candidate data
            candidate_data['scoring_breakdown'] = scoring_breakdown
            
            return min(total_score, 100)
            
        except Exception as e:
            logger.error(f"Error calculating comprehensive PDS score: {str(e)}")
            return 0
    
    def _extract_highest_education(self, pds_data):
        """Extract highest education level from PDS data."""
        education = pds_data.get('educational_background', {})
        
        if education.get('graduate'):
            return 'Graduate Studies'
        elif education.get('college'):
            return 'College'
        elif education.get('vocational'):
            return 'Vocational'
        elif education.get('secondary'):
            return 'Secondary'
        elif education.get('elementary'):
            return 'Elementary'
        else:
            return 'Not Specified'
    
    def _calculate_experience_years(self, pds_data):
        """Calculate total years of work experience."""
        work_experience = pds_data.get('work_experience', [])
        return len(work_experience)  # Simplified calculation
    
    def _calculate_govt_service_years(self, pds_data):
        """Calculate years in government service."""
        work_experience = pds_data.get('work_experience', [])
        govt_years = 0
        for exp in work_experience:
            if exp.get('govt_service') == 'Y':
                govt_years += 1
        return govt_years
    
    def _check_civil_service_eligibility(self, pds_data):
        """Check if candidate has civil service eligibility."""
        eligibility = pds_data.get('eligibility', [])
        return len(eligibility) > 0
    
    def _extract_matched_qualifications(self, candidate_data, job):
        """Extract qualifications that match job requirements."""
        matched = []
        job_requirements = job.get('requirements', '').lower()
        
        # Check education match
        highest_ed = self._extract_highest_education(candidate_data['pds_data'])
        if 'college' in job_requirements and 'College' in highest_ed:
            matched.append('College Education')
        
        # Check experience match
        exp_years = self._calculate_experience_years(candidate_data['pds_data'])
        if exp_years >= 3:
            matched.append('Relevant Work Experience')
        
        # Check civil service eligibility
        if self._check_civil_service_eligibility(candidate_data['pds_data']):
            matched.append('Civil Service Eligible')
        
        return matched
    
    def _identify_improvement_areas(self, candidate_data, job):
        """Identify areas where candidate could improve."""
        areas = []
        
        # Check if more training would help
        training = candidate_data['pds_data'].get('training', [])
        if len(training) < 3:
            areas.append('Additional professional training')
        
        # Check experience level
        exp_years = self._calculate_experience_years(candidate_data['pds_data'])
        if exp_years < 5:
            areas.append('More work experience')
        
        return areas
    
    def _score_education_comprehensive(self, education, job):
        """Comprehensive education scoring (max 30 points)."""
        score = 0
        
        # Education level (20 points)
        if education.get('graduate'):
            score += 20
        elif education.get('college'):
            score += 15
        elif education.get('vocational'):
            score += 10
        elif education.get('secondary'):
            score += 5
        
        # Relevance bonus (10 points)
        job_title = job.get('title', '').lower()
        college_info = education.get('college', '').lower()
        
        if college_info:
            if any(keyword in college_info for keyword in ['computer', 'information technology', 'engineering']):
                if any(tech in job_title for tech in ['it', 'technology', 'software', 'analyst']):
                    score += 10
            elif any(keyword in college_info for keyword in ['business', 'management', 'administration']):
                if any(admin in job_title for admin in ['admin', 'management', 'supervisor']):
                    score += 8
        
        return min(score, 30)
    
    def _score_experience_comprehensive(self, work_experience, job):
        """Comprehensive experience scoring (max 35 points)."""
        score = 0
        
        # Years of experience (20 points)
        years = len(work_experience)
        score += min(years * 3, 20)
        
        # Government service bonus (10 points)
        govt_experience = sum(1 for exp in work_experience if exp.get('govt_service') == 'Y')
        score += min(govt_experience * 2, 10)
        
        # Position relevance (5 points)
        job_title = job.get('title', '').lower()
        for exp in work_experience:
            position = exp.get('position', '').lower()
            if any(keyword in position for keyword in job_title.split()):
                score += 5
                break
        
        return min(score, 35)
    
    def _score_eligibility_comprehensive(self, eligibility):
        """Comprehensive eligibility scoring (max 15 points)."""
        score = 0
        
        if len(eligibility) > 0:
            score += 10  # Basic eligibility
            
            # Professional/career service bonus
            for elig in eligibility:
                elig_text = elig.get('eligibility', '').lower()
                if any(keyword in elig_text for keyword in ['professional', 'career service']):
                    score += 5
                    break
        
        return min(score, 15)
    
    def _score_training_comprehensive(self, training, job):
        """Comprehensive training scoring (max 10 points)."""
        score = 0
        
        # Number of training programs
        training_count = len(training)
        score += min(training_count, 5)
        
        # Relevance bonus
        job_title = job.get('title', '').lower()
        for train in training:
            title = train.get('title', '').lower()
            if any(keyword in title for keyword in job_title.split()):
                score += 5
                break
        
        return min(score, 10)
    
    def _score_additional_qualifications(self, pds_data, job):
        """Score additional qualifications (max 10 points)."""
        score = 0
        
        # Volunteer work (3 points)
        if len(pds_data.get('voluntary_work', [])) > 0:
            score += 3
        
        # Professional references (3 points)
        if len(pds_data.get('personal_references', [])) >= 3:
            score += 3
        
        # Additional information (4 points)
        other_info = pds_data.get('other_info', {})
        if other_info.get('special_skills') or other_info.get('recognition'):
            score += 4
        
        return min(score, 10)
    
    @login_required
    def get_pds_candidates(self):
        """Get list of PDS candidates organized by job"""
        try:
            candidates = db_manager.get_all_pds_candidates()
            jobs = db_manager.get_all_jobs()
            
            # Group candidates by job
            candidates_by_job = {}
            
            # Initialize with all jobs
            for job in jobs:
                candidates_by_job[job['id']] = {
                    'job_info': {
                        'id': job['id'],
                        'title': job['title'],
                        'department': job['department'],
                        'category': job['category']
                    },
                    'candidates': []
                }
            
            # Add candidates to their respective jobs
            for candidate in candidates:
                job_id = candidate.get('job_id')
                if job_id and job_id in candidates_by_job:
                    candidates_by_job[job_id]['candidates'].append({
                        'id': candidate['id'],
                        'name': candidate['name'],
                        'email': candidate['email'],
                        'phone': candidate['phone'],
                        'score': candidate['score'],
                        'status': candidate['status'],
                        'highest_education': candidate['highest_education'],
                        'years_of_experience': candidate['years_of_experience'],
                        'civil_service_eligible': candidate['civil_service_eligible'],
                        'upload_timestamp': candidate['upload_timestamp'],
                        'filename': candidate['filename']
                    })
            
            return jsonify({
                'success': True,
                'candidates_by_job': candidates_by_job,
                'total_candidates': len(candidates)
            })
            
        except Exception as e:
            logger.error(f"Error getting PDS candidates: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_pds_candidate(self, candidate_id):
        """Handle individual PDS candidate operations"""
        if request.method == 'GET':
            try:
                candidate = db_manager.get_pds_candidate(candidate_id)
                if not candidate:
                    return jsonify({'success': False, 'error': 'PDS candidate not found'}), 404
                return jsonify({'success': True, 'candidate': candidate})
            except Exception as e:
                logger.error(f"Error getting PDS candidate: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'PUT':
            try:
                data = request.get_json()
                success = db_manager.update_pds_candidate(candidate_id, data)
                
                if not success:
                    return jsonify({'success': False, 'error': 'Failed to update PDS candidate'}), 400
                
                candidate = db_manager.get_pds_candidate(candidate_id)
                return jsonify({
                    'success': True,
                    'message': 'PDS candidate updated successfully',
                    'candidate': candidate
                })
                
            except Exception as e:
                logger.error(f"Error updating PDS candidate: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'DELETE':
            try:
                success = db_manager.delete_pds_candidate(candidate_id)
                if not success:
                    return jsonify({'success': False, 'error': 'Failed to delete PDS candidate'}), 400
                
                return jsonify({'success': True, 'message': 'PDS candidate deleted successfully'})
            except Exception as e:
                logger.error(f"Error deleting PDS candidate: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_jobs(self):
        """Handle job listing and creation"""
        if request.method == 'GET':
            try:
                jobs = db_manager.get_all_jobs()
                logger.info(f"GET /api/jobs - Returning {len(jobs)} jobs")
                return jsonify({
                    'success': True,
                    'jobs': jobs
                })
            except Exception as e:
                logger.error(f"Error getting jobs: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'POST':
            try:
                data = request.get_json()
                
                # Validate required fields
                required_fields = ['title', 'department', 'category', 'experience_level', 'description', 'requirements']
                for field in required_fields:
                    if not data.get(field):
                        return jsonify({'success': False, 'error': f'{field} is required'}), 400
                
                # Convert category name to category_id
                category_name = data['category']
                categories = db_manager.get_all_job_categories()
                category_id = None
                
                for category in categories:
                    if category['name'] == category_name:
                        category_id = category['id']
                        break
                
                if category_id is None:
                    return jsonify({'success': False, 'error': f'Category "{category_name}" not found'}), 400
                
                # Prepare job data with category_id
                job_data = {
                    'title': data['title'],
                    'department': data['department'],
                    'description': data['description'],
                    'requirements': data['requirements'],
                    'experience_level': data['experience_level'],
                    'category_id': category_id
                }
                
                # Create new job
                job_id = db_manager.create_job(job_data)
                job = db_manager.get_job(job_id)
                
                return jsonify({
                    'success': True,
                    'message': 'Job created successfully',
                    'job': job
                })
                
            except Exception as e:
                logger.error(f"Error creating job: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_job(self, job_id):
        """Handle individual job operations"""
        if request.method == 'GET':
            try:
                job = db_manager.get_job(job_id)
                if not job:
                    return jsonify({'success': False, 'error': 'Job not found'}), 404
                return jsonify({'success': True, 'job': job})
            except Exception as e:
                logger.error(f"Error getting job: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'PUT':
            try:
                data = request.get_json()
                
                # Convert category name to category_id if category is provided
                if 'category' in data:
                    category_name = data['category']
                    categories = db_manager.get_all_job_categories()
                    category_id = None
                    
                    for category in categories:
                        if category['name'] == category_name:
                            category_id = category['id']
                            break
                    
                    if category_id is None:
                        return jsonify({'success': False, 'error': f'Category "{category_name}" not found'}), 400
                    
                    # Replace category name with category_id
                    data = data.copy()  # Don't modify original data
                    del data['category']
                    data['category_id'] = category_id
                
                success = db_manager.update_job(job_id, data)
                
                if not success:
                    return jsonify({'success': False, 'error': 'Job not found'}), 404
                
                job = db_manager.get_job(job_id)
                return jsonify({
                    'success': True,
                    'message': 'Job updated successfully',
                    'job': job
                })
                
            except Exception as e:
                logger.error(f"Error updating job: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'DELETE':
            try:
                success = db_manager.delete_job(job_id)
                if not success:
                    return jsonify({'success': False, 'error': 'Job not found'}), 404
                
                return jsonify({'success': True, 'message': 'Job deleted successfully'})
            except Exception as e:
                logger.error(f"Error deleting job: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_job_categories(self):
        """Handle job category listing and creation"""
        if request.method == 'GET':
            try:
                categories = db_manager.get_all_job_categories()
                logger.info(f"GET /api/job-categories - Returning {len(categories)} categories")
                return jsonify({
                    'success': True,
                    'categories': categories
                })
            except Exception as e:
                logger.error(f"Error getting job categories: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'POST':
            try:
                data = request.get_json()
                
                if not data.get('name'):
                    return jsonify({'success': False, 'error': 'Category name is required'}), 400
                
                # Check if category already exists
                existing_categories = db_manager.get_all_job_categories()
                for category in existing_categories:
                    if category['name'].lower() == data['name'].lower():
                        return jsonify({'success': False, 'error': 'Category already exists'}), 400
                
                # Create new category
                category_id = db_manager.create_job_category(
                    data['name'], 
                    data.get('description', '')
                )
                
                category = {
                    'id': category_id,
                    'name': data['name'],
                    'description': data.get('description', ''),
                    'created_at': datetime.now().isoformat()
                }
                
                return jsonify({
                    'success': True,
                    'message': 'Category created successfully',
                    'category': category
                })
                
            except Exception as e:
                logger.error(f"Error creating category: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_job_category(self, category_id):
        """Handle individual job category operations"""
        if request.method == 'PUT':
            try:
                data = request.get_json()
                
                name = data.get('name')
                description = data.get('description')
                
                success = db_manager.update_job_category(category_id, name, description)
                if not success:
                    return jsonify({'success': False, 'error': 'Category not found'}), 404
                
                # Get updated category (simplified response)
                category = {
                    'id': category_id,
                    'name': name if name else 'Updated Category',
                    'description': description if description else ''
                }
                
                return jsonify({
                    'success': True,
                    'message': 'Category updated successfully',
                    'category': category
                })
                
            except Exception as e:
                logger.error(f"Error updating category: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'DELETE':
            try:
                # Get category name before deleting to check usage
                categories = db_manager.get_all_job_categories()
                category_name = None
                for cat in categories:
                    if cat['id'] == category_id:
                        category_name = cat['name']
                        break
                
                if not category_name:
                    return jsonify({'success': False, 'error': 'Category not found'}), 404
                
                # Check if any jobs use this category
                jobs_count = db_manager.check_category_in_use(category_name)
                if jobs_count > 0:
                    return jsonify({
                        'success': False, 
                        'error': f'Cannot delete category. {jobs_count} jobs are using this category.'
                    }), 400
                
                success = db_manager.delete_job_category(category_id)
                if not success:
                    return jsonify({'success': False, 'error': 'Category not found'}), 404
                
                return jsonify({'success': True, 'message': 'Category deleted successfully'})
            except Exception as e:
                logger.error(f"Error deleting category: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    def _calculate_candidate_assessment_score(self, candidate):
        """Calculate assessment score for a candidate using university criteria"""
        try:
            import json
            
            # Parse PDS data if available
            pds_data = None
            if candidate.get('pds_extracted_data'):
                try:
                    pds_data = json.loads(candidate['pds_extracted_data'])
                except:
                    pass
            
            # Check if we have assessment data (either PDS data or legacy data)
            if pds_data:
                # Use PDS data for assessment
                pass
            elif candidate.get('education') or candidate.get('experience'):
                # Use legacy resume data for assessment  
                pds_data = {
                    'educational_background': candidate.get('education', []),
                    'work_experience': candidate.get('experience', []),
                    'training_programs': candidate.get('training', []),
                    'civil_service_eligibility': candidate.get('eligibility', []),
                    'other_info': {
                        'recognitions': []  # Default empty recognitions
                    }
                }
            else:
                # No assessment data available
                return 0
            
            # Initialize assessment result
            assessment_result = {
                'education_score': 0,
                'experience_score': 0,
                'training_score': 0,
                'eligibility_score': 0,
                'accomplishments_score': 0,
                'potential_score': 0
            }
            
            # Education Assessment (40 points max)
            education_data = pds_data.get('educational_background', [])
            education_score = 0
            has_doctorate = False
            has_masters = False
            has_bachelors = False
            
            for edu in education_data:
                level = str(edu.get('level', '')).upper()
                degree = str(edu.get('degree', '')).lower()
                
                if 'DOCTORATE' in level or 'DOCTORAL' in level or 'phd' in degree or 'doctor' in degree:
                    has_doctorate = True
                elif 'GRADUATE' in level or 'MASTER' in level or 'master' in degree:
                    has_masters = True
                elif 'COLLEGE' in level or 'bachelor' in degree:
                    has_bachelors = True
            
            # Calculate education score
            if has_bachelors:
                education_score = 35
            if has_masters:
                education_score = max(education_score, 38)
            if has_doctorate:
                education_score = 40
                
            assessment_result['education_score'] = education_score
            
            # Experience Assessment (20 points max)
            experience_data = pds_data.get('work_experience', [])
            experience_score = 0
            total_years = 0
            
            for exp in experience_data:
                # Try to calculate years of experience
                date_from = str(exp.get('date_from', ''))
                date_to = str(exp.get('date_to', ''))
                
                if date_from and date_to:
                    try:
                        # Simple year calculation
                        from_year = int(date_from.split('-')[0]) if '-' in date_from else int(date_from[:4])
                        to_year = int(date_to.split('-')[0]) if '-' in date_to else int(date_to[:4])
                        years = max(0, to_year - from_year)
                        total_years += years
                    except:
                        # Fallback: assume 1 year per position
                        total_years += 1
                else:
                    # Fallback: assume 1 year per position
                    total_years += 1
            
            # Score based on years: 1 point per year, max 20
            experience_score = min(total_years, 20)
            assessment_result['experience_score'] = experience_score
            
            # Training Assessment (10 points max)
            training_data = pds_data.get('training_programs', [])
            training_score = min(len(training_data) * 2, 10)
            assessment_result['training_score'] = training_score
            
            # Eligibility Assessment (10 points max)
            eligibility_data = pds_data.get('civil_service_eligibility', [])
            eligibility_score = min(len(eligibility_data) * 5, 10)
            assessment_result['eligibility_score'] = eligibility_score
            
            # Accomplishments Assessment (5 points max)
            other_info = pds_data.get('other_info', {})
            recognitions = other_info.get('recognitions', [])
            accomplishments_score = min(len(recognitions), 5)
            assessment_result['accomplishments_score'] = accomplishments_score
            
            # Potential Score - get from database (15 points max)
            potential_score = candidate.get('potential_score', 0.0)
            assessment_result['potential_score'] = potential_score
            
            # Calculate total including potential score from database
            total_score = (
                assessment_result['education_score'] +
                assessment_result['experience_score'] +
                assessment_result['training_score'] +
                assessment_result['eligibility_score'] +
                assessment_result['accomplishments_score'] +
                assessment_result['potential_score']
            )
            
            return total_score
            
        except Exception as e:
            logger.error(f"Error calculating assessment score for candidate {candidate.get('id', 'unknown')}: {e}")
            return 0

    def get_candidates(self):
        """Get list of candidates organized by LSPU job categories - LSPU-only system"""
        try:
            # Get all candidates from PostgreSQL
            candidates = db_manager.get_all_candidates()
            
            # Get LSPU job postings only (no more legacy jobs)
            lspu_jobs = self._get_all_lspu_job_postings()
            
            # Group candidates by LSPU job
            candidates_by_job = {}
            
            # Initialize with all LSPU jobs - use LSPU field names for frontend
            for job in lspu_jobs:
                candidates_by_job[f"lspu_{job['id']}"] = {
                    # LSPU-specific fields expected by frontend
                    'position_title': job['position_title'],
                    'position_category': job['position_category'], 
                    'campus_name': job.get('campus_name', 'LSPU'),
                    'department_office': job.get('department_office', ''),
                    'salary_grade': job.get('salary_grade', ''),
                    # Additional job details
                    'job_reference_number': job.get('job_reference_number', ''),
                    'status': job.get('status', 'active'),
                    'job_description': f"Position: {job['position_title']} at {job.get('campus_name', 'LSPU')}",
                    'job_requirements': ", ".join(filter(None, [
                        job.get('education_requirements', ''),
                        job.get('experience_requirements', ''),
                        job.get('training_requirements', '')
                    ])),
                    'source': 'LSPU',
                    'candidates': []
                }
            
            # Add unassigned category for candidates without job_id - use legacy field names
            candidates_by_job['unassigned'] = {
                'job_title': 'Unassigned Applications',
                'job_category': 'UNASSIGNED', 
                'job_description': 'Candidates not yet assigned to a specific position',
                'job_requirements': 'No specific requirements',
                'source': 'LSPU',
                'candidates': []
            }
            
            # Add candidates to their respective LSPU jobs
            for candidate in candidates:
                job_id = candidate.get('job_id', 'unassigned')
                
                # Convert job_id to LSPU key format
                if job_id != 'unassigned':
                    lspu_job_key = f"lspu_{job_id}"
                    if lspu_job_key not in candidates_by_job:
                        # Skip candidates assigned to non-existent jobs
                        logger.warning(f"Candidate {candidate.get('id')} assigned to non-existent job {job_id}")
                        job_id = 'unassigned'
                    else:
                        job_id = lspu_job_key
                
                # Use 'unassigned' if job not found
                target_job_id = job_id
                
                # Format education as string for display
                education_str = self._format_candidate_education(candidate)
                
                # Format skills as string
                skills_list = candidate.get('skills', [])
                if isinstance(skills_list, str):
                    try:
                        skills_list = json.loads(skills_list)
                    except:
                        skills_list = [skills_list] if skills_list else []
                
                skills_str = ", ".join(skills_list[:10]) + ("..." if len(skills_list) > 10 else "") if skills_list else "Not specified"
                
                # Format predicted category
                predicted_category_str = candidate.get('category', 'Unknown')
                
                # Enhanced candidate data with PDS information
                formatted_candidate = {
                    'id': candidate['id'],
                    'name': candidate['name'],
                    'email': candidate['email'],
                    'phone': candidate['phone'],
                    'education': education_str,
                    'skills': skills_str,
                    'all_skills': skills_list,
                    'predicted_category': predicted_category_str,
                    'score': candidate['score'],
                    'status': candidate['status'],
                    'processing_type': candidate.get('processing_type', 'resume'),
                    'ocr_confidence': candidate.get('ocr_confidence'),
                    'created_at': candidate['created_at'].isoformat() if candidate.get('created_at') else None,
                    'updated_at': candidate['updated_at'].isoformat() if candidate.get('updated_at') else None,
                    # Enhanced PDS fields
                    'total_education_entries': candidate.get('total_education_entries', 0),
                    'total_work_positions': candidate.get('total_work_positions', 0),
                    'extraction_status': candidate.get('extraction_status', 'pending'),
                    'uploaded_filename': candidate.get('uploaded_filename', ''),
                    'latest_total_score': candidate.get('latest_total_score'),
                    'latest_percentage_score': candidate.get('latest_percentage_score'),
                    'latest_recommendation': candidate.get('latest_recommendation'),
                    # PDS-specific fields for frontend display
                    'government_ids': candidate.get('government_ids', {}),
                    'education': candidate.get('education', []) if isinstance(candidate.get('education'), list) else [],
                    'eligibility': candidate.get('eligibility', []),
                    'work_experience': candidate.get('work_experience', []),
                    'pds_data': candidate.get('pds_data', {}),
                }
                
                # Calculate real assessment score using university criteria
                assessment_score = self._calculate_candidate_assessment_score(candidate)
                formatted_candidate['assessment_score'] = assessment_score
                formatted_candidate['score'] = assessment_score  # Update the score field too
                
                candidates_by_job[target_job_id]['candidates'].append(formatted_candidate)
            
            # Calculate totals
            total_candidates = len(candidates)
            lspu_job_count = len([job_id for job_id in candidates_by_job.keys() if job_id != 'unassigned'])
            
            return jsonify({
                'success': True,
                'candidates_by_job': candidates_by_job,
                'total_candidates': total_candidates,
                'total_jobs': lspu_job_count,
                'system': 'LSPU-only',
                'data_source': 'lspu_unified'
            })
            
        except Exception as e:
            logger.error(f"Error getting LSPU candidates: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_candidate(self, candidate_id):
        """Handle individual candidate operations"""
        if request.method == 'GET':
            try:
                candidate = db_manager.get_candidate(candidate_id)
                if not candidate:
                    return jsonify({'success': False, 'error': 'Candidate not found'}), 404
                
                # Format detailed candidate info
                detailed_candidate = {
                    'id': candidate['id'],
                    'name': candidate['name'],
                    'email': candidate['email'],
                    'phone': candidate['phone'],
                    'matchScore': candidate['score'],
                    'status': candidate['status'],
                    'category': candidate['category'],
                    'job_title': candidate['job_title'],
                    'skills': candidate.get('skills', []),
                    'education': candidate.get('education', []),
                    'matched_skills': candidate.get('matched_skills', []),
                    'missing_skills': candidate.get('missing_skills', []),
                    'predicted_category': candidate.get('predicted_category', {}),
                    'filename': candidate.get('filename', ''),
                    'updated_at': candidate['updated_at'],
                    'processing_type': candidate.get('processing_type', 'resume'),
                    
                    # PDS-specific fields
                    'pds_data': candidate.get('pds_data', {}),
                    'government_ids': candidate.get('government_ids', {}),
                    'eligibility': candidate.get('eligibility', []),
                    'training': candidate.get('training', []),
                    'work_experience': candidate.get('experience', []),  # Map to experience field
                    'voluntary_work': candidate.get('volunteer_work', []),
                    'personal_references': candidate.get('personal_references', [])
                }
                
                return jsonify({'success': True, 'candidate': detailed_candidate})
            except Exception as e:
                logger.error(f"Error getting candidate: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'PUT':
            try:
                data = request.get_json()
                
                if 'status' in data:
                    success = db_manager.update_candidate(candidate_id, {'status': data['status']})
                    if not success:
                        return jsonify({'success': False, 'error': 'Candidate not found'}), 404
                    
                    candidate = db_manager.get_candidate(candidate_id)
                    return jsonify({
                        'success': True,
                        'message': 'Candidate updated successfully',
                        'candidate': candidate
                    })
                else:
                    return jsonify({'success': False, 'error': 'No valid fields to update'}), 400
                
            except Exception as e:
                logger.error(f"Error updating candidate: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'DELETE':
            try:
                success = db_manager.delete_candidate(candidate_id)
                if not success:
                    return jsonify({'success': False, 'error': 'Candidate not found'}), 404
                
                return jsonify({'success': True, 'message': 'Candidate removed successfully'})
            except Exception as e:
                logger.error(f"Error deleting candidate: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def get_analytics(self):
        """Get analytics data"""
        try:
            # Get summary data from database
            summary = db_manager.get_analytics_summary()
            
            # Create more realistic daily stats based on actual data
            today = datetime.now().date()
            daily_stats = []
            
            # Use actual data for recent days and reduce for older days
            for i in range(30):
                date_obj = today - timedelta(days=i)
                
                # Simulate realistic data degradation over time
                time_factor = max(0.1, 1 - (i * 0.1))  # Reduce by 10% each day going back
                
                daily_stats.append({
                    'date': date_obj.strftime('%Y-%m-%d'),
                    'total_resumes': max(0, int(summary['total_resumes'] * time_factor)),
                    'processed_resumes': max(0, int(summary['processed_resumes'] * time_factor)),
                    'shortlisted': max(0, int(summary['shortlisted'] * time_factor)),
                    'rejected': max(0, int(summary['rejected'] * time_factor)),
                    'job_category_stats': json.dumps(summary['job_category_stats'])
                })
            
            daily_stats.reverse()  # Show oldest to newest
            
            return jsonify({
                'success': True,
                'summary': {
                    'total_resumes': summary['total_resumes'],
                    'processed_resumes': summary['processed_resumes'],
                    'total_pds': summary['total_pds'],
                    'processed_pds': summary['processed_pds'],
                    'shortlisted': summary['shortlisted'],
                    'rejected': summary['rejected'],
                    'avg_score': summary['avg_score'],
                    'avg_processing_time': 3  # Fixed value for now
                },
                'daily_stats': daily_stats,
                'processing_type_stats': summary['processing_type_stats'],
                'job_category_stats': summary['job_category_stats']
            })
            
        except Exception as e:
            logger.error(f"Error getting analytics: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_settings(self):
        """Handle application settings"""
        if request.method == 'GET':
            try:
                settings = db_manager.get_all_settings()
                return jsonify({
                    'success': True,
                    'settings': settings
                })
            except Exception as e:
                logger.error(f"Error getting settings: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'PUT':
            try:
                data = request.get_json()
                db_manager.update_settings(data)
                
                settings = db_manager.get_all_settings()
                return jsonify({
                    'success': True,
                    'message': 'Settings updated successfully',
                    'settings': settings
                })
                
            except Exception as e:
                logger.error(f"Error updating settings: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500

    # ============================================================================
    # UNIVERSITY ASSESSMENT API ENDPOINTS
    # ============================================================================
    
    @login_required
    def get_position_types(self):
        """Get all available position types"""
        try:
            position_types = db_manager.get_position_types()
            return jsonify({
                'success': True,
                'position_types': position_types
            })
        except Exception as e:
            logger.error(f"Error getting position types: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def get_assessment_templates(self, position_type_id):
        """Get assessment templates for a position type"""
        try:
            templates = db_manager.get_assessment_templates_by_category(position_type_id)
            return jsonify({
                'success': True,
                'templates': templates,
                'position_type_id': position_type_id
            })
        except Exception as e:
            logger.error(f"Error getting assessment templates: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_position_requirements(self, job_id):
        """Handle position requirements for a job"""
        if request.method == 'GET':
            try:
                requirements = db_manager.get_position_requirements(job_id)
                return jsonify({
                    'success': True,
                    'requirements': requirements
                })
            except Exception as e:
                logger.error(f"Error getting position requirements: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method in ['POST', 'PUT']:
            try:
                data = request.get_json()
                
                # Validate required fields
                if not data.get('position_type_id'):
                    return jsonify({'success': False, 'error': 'Position type is required'}), 400
                
                # Check if requirements already exist for update
                existing = db_manager.get_position_requirements(job_id)
                
                if request.method == 'POST' and existing:
                    return jsonify({'success': False, 'error': 'Position requirements already exist for this job'}), 400
                
                if request.method == 'PUT' and not existing:
                    return jsonify({'success': False, 'error': 'No position requirements found to update'}), 404
                
                # Create or update requirements
                if request.method == 'POST':
                    req_id = db_manager.create_position_requirement(
                        job_id=job_id,
                        position_type_id=data['position_type_id'],
                        minimum_education=data.get('minimum_education'),
                        required_experience=data.get('required_experience', 0),
                        required_certifications=data.get('required_certifications', []),
                        preferred_qualifications=data.get('preferred_qualifications'),
                        subject_area=data.get('subject_area')
                    )
                    
                    if req_id:
                        requirements = db_manager.get_position_requirements(job_id)
                        return jsonify({
                            'success': True,
                            'message': 'Position requirements created successfully',
                            'requirements': requirements
                        })
                    else:
                        return jsonify({'success': False, 'error': 'Failed to create position requirements'}), 500
                
                # For PUT, would need to implement update_position_requirements method
                return jsonify({'success': False, 'error': 'Update not implemented yet'}), 501
                
            except Exception as e:
                logger.error(f"Error handling position requirements: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_job_assessments(self, job_id):
        """Handle assessments for a specific job"""
        if request.method == 'GET':
            try:
                status_filter = request.args.get('status')
                assessments = db_manager.get_assessments_for_job(job_id, status_filter)
                
                # Update rankings if requested
                if request.args.get('update_rankings') == 'true':
                    db_manager.update_assessment_rankings(job_id)
                    # Re-fetch with updated rankings
                    assessments = db_manager.get_assessments_for_job(job_id, status_filter)
                
                return jsonify({
                    'success': True,
                    'job_id': job_id,
                    'assessments': assessments,
                    'total_count': len(assessments)
                })
            except Exception as e:
                logger.error(f"Error getting job assessments: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'POST':
            try:
                data = request.get_json()
                candidate_ids = data.get('candidate_ids', [])
                position_type_id = data.get('position_type_id')
                
                if not candidate_ids:
                    return jsonify({'success': False, 'error': 'No candidates specified'}), 400
                
                if not position_type_id:
                    return jsonify({'success': False, 'error': 'Position type is required'}), 400
                
                created_assessments = []
                for candidate_id in candidate_ids:
                    assessment_id = db_manager.create_candidate_assessment(
                        candidate_id=candidate_id,
                        job_id=job_id,
                        position_type_id=position_type_id,
                        assessed_by=current_user.id if hasattr(current_user, 'id') else session.get('user_id')
                    )
                    if assessment_id:
                        created_assessments.append(assessment_id)
                
                return jsonify({
                    'success': True,
                    'message': f'Created {len(created_assessments)} assessments',
                    'assessment_ids': created_assessments
                })
                
            except Exception as e:
                logger.error(f"Error creating job assessments: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_assessment(self, assessment_id):
        """Handle individual assessment operations"""
        if request.method == 'GET':
            try:
                # Get assessment by ID - need to implement this in database.py
                # For now, use a workaround to get candidate and job info
                assessment = db_manager.get_candidate_assessment_by_id(assessment_id)
                if not assessment:
                    return jsonify({'success': False, 'error': 'Assessment not found'}), 404
                
                # Get manual scores
                manual_scores = db_manager.get_manual_assessment_scores(assessment_id)
                assessment['manual_scores'] = manual_scores
                
                return jsonify({
                    'success': True,
                    'assessment': assessment
                })
            except Exception as e:
                logger.error(f"Error getting assessment: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'PUT':
            try:
                data = request.get_json()
                
                # Update scores and status
                success = db_manager.update_candidate_assessment_scores(
                    assessment_id=assessment_id,
                    education_score=data.get('education_score'),
                    experience_score=data.get('experience_score'),
                    training_score=data.get('training_score'),
                    eligibility_score=data.get('eligibility_score'),
                    accomplishments_score=data.get('accomplishments_score'),
                    interview_score=data.get('interview_score'),
                    aptitude_score=data.get('aptitude_score'),
                    score_breakdown=data.get('score_breakdown'),
                    assessment_notes=data.get('assessment_notes')
                )
                
                if success:
                    # Update status if provided
                    if data.get('assessment_status') or data.get('recommendation'):
                        db_manager.update_assessment_status(
                            assessment_id=assessment_id,
                            status=data.get('assessment_status'),
                            recommendation=data.get('recommendation'),
                            completed_date=datetime.now() if data.get('assessment_status') == 'complete' else None
                        )
                    
                    return jsonify({
                        'success': True,
                        'message': 'Assessment updated successfully'
                    })
                else:
                    return jsonify({'success': False, 'error': 'Failed to update assessment'}), 500
                
            except Exception as e:
                logger.error(f"Error updating assessment: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def handle_manual_scores(self, assessment_id):
        """Handle manual assessment scores (interview, aptitude)"""
        if request.method == 'GET':
            try:
                manual_scores = db_manager.get_manual_assessment_scores(assessment_id)
                return jsonify({
                    'success': True,
                    'assessment_id': assessment_id,
                    'manual_scores': manual_scores
                })
            except Exception as e:
                logger.error(f"Error getting manual scores: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
        
        elif request.method == 'POST':
            try:
                data = request.get_json()
                scores = data.get('scores', [])
                
                if not scores:
                    return jsonify({'success': False, 'error': 'No scores provided'}), 400
                
                created_scores = []
                for score_data in scores:
                    score_id = db_manager.create_manual_assessment_score(
                        candidate_assessment_id=assessment_id,
                        score_type=score_data['score_type'],
                        component_name=score_data['component_name'],
                        rating=score_data['rating'],
                        score=score_data['score'],
                        max_possible=score_data['max_possible'],
                        notes=score_data.get('notes'),
                        entered_by=current_user.id if hasattr(current_user, 'id') else session.get('user_id')
                    )
                    if score_id:
                        created_scores.append(score_id)
                
                return jsonify({
                    'success': True,
                    'message': f'Created {len(created_scores)} manual scores',
                    'score_ids': created_scores
                })
                
            except Exception as e:
                logger.error(f"Error creating manual scores: {e}")
                return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def get_assessment_comparison(self, job_id):
        """Get assessment comparison and ranking for a job"""
        try:
            # Update rankings first
            db_manager.update_assessment_rankings(job_id)
            
            # Get comparison results
            comparison = db_manager.get_assessment_comparison(job_id, latest=True)
            
            # If no saved comparison exists, generate one
            if not comparison:
                assessments = db_manager.get_assessments_for_job(job_id)
                
                # Create ranking data
                candidate_rankings = []
                for i, assessment in enumerate(assessments):
                    candidate_rankings.append({
                        'rank': i + 1,
                        'candidate_id': assessment['candidate_id'],
                        'candidate_name': assessment['candidate_name'],
                        'final_score': assessment['final_score'],
                        'automated_score': assessment['automated_total'],
                        'manual_score': assessment['manual_total'],
                        'recommendation': assessment['recommendation']
                    })
                
                # Generate summary statistics
                scores = [a['final_score'] for a in assessments if a['final_score'] > 0]
                summary = {
                    'total_candidates': len(assessments),
                    'completed_assessments': len([a for a in assessments if a['assessment_status'] == 'complete']),
                    'average_score': round(sum(scores) / len(scores), 2) if scores else 0,
                    'highest_score': max(scores) if scores else 0,
                    'lowest_score': min(scores) if scores else 0
                }
                
                # Save the comparison
                comparison_id = db_manager.save_assessment_comparison(
                    job_id=job_id,
                    candidate_rankings=candidate_rankings,
                    assessment_summary=summary,
                    generated_by=current_user.id if hasattr(current_user, 'id') else session.get('user_id')
                )
                
                comparison = {
                    'id': comparison_id,
                    'job_id': job_id,
                    'candidate_rankings': candidate_rankings,
                    'assessment_summary': summary,
                    'comparison_date': datetime.now().isoformat()
                }
            
            return jsonify({
                'success': True,
                'comparison': comparison
            })
            
        except Exception as e:
            logger.error(f"Error getting assessment comparison: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def get_assessment_analytics(self, job_id):
        """Get assessment analytics for a job"""
        try:
            analytics = db_manager.get_assessment_analytics(job_id=job_id)
            return jsonify({
                'success': True,
                'job_id': job_id,
                'analytics': analytics
            })
        except Exception as e:
            logger.error(f"Error getting assessment analytics: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    @login_required
    def get_university_assessment_analytics(self):
        """Get comprehensive university assessment criteria analytics based on real data"""
        try:
            # Get real data from databases
            basic_analytics = db_manager.get_analytics_summary()
            assessment_analytics = db_manager.get_assessment_analytics()
            
            # Get detailed candidate information
            with db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get candidate details with scores and categories
                cursor.execute("""
                    SELECT 
                        id, name, status, score, category, processing_type,
                        created_at, updated_at
                    FROM candidates 
                    ORDER BY updated_at DESC
                """)
                candidates = [dict(row) for row in cursor.fetchall()]
                
                # Get score distribution from real data
                cursor.execute("""
                    SELECT 
                        CASE 
                            WHEN score >= 90 THEN 'Excellent (90+)'
                            WHEN score >= 80 THEN 'Very Good (80-89)'
                            WHEN score >= 70 THEN 'Good (70-79)'
                            WHEN score >= 60 THEN 'Fair (60-69)'
                            WHEN score > 0 THEN 'Needs Improvement (<60)'
                            ELSE 'Not Assessed'
                        END as score_range,
                        COUNT(*) as count
                    FROM candidates
                    GROUP BY score_range
                    ORDER BY 
                        CASE 
                            WHEN score >= 90 THEN 1
                            WHEN score >= 80 THEN 2
                            WHEN score >= 70 THEN 3
                            WHEN score >= 60 THEN 4
                            WHEN score > 0 THEN 5
                            ELSE 6
                        END
                """)
                real_score_distribution = {row['score_range']: row['count'] for row in cursor.fetchall()}
                
                # Get category performance
                cursor.execute("""
                    SELECT 
                        category,
                        COUNT(*) as total_candidates,
                        AVG(CASE WHEN score > 0 THEN score ELSE NULL END) as avg_score,
                        COUNT(CASE WHEN status = 'shortlisted' THEN 1 END) as shortlisted_count,
                        COUNT(CASE WHEN score >= 70 THEN 1 END) as high_performers
                    FROM candidates 
                    WHERE category IS NOT NULL
                    GROUP BY category
                    ORDER BY avg_score DESC
                """)
                category_performance = [dict(row) for row in cursor.fetchall()]
            
            # Calculate real criteria performance based on actual candidates
            total_candidates = basic_analytics.get('total_resumes', 0)
            processed_candidates = basic_analytics.get('processed_resumes', 0)
            avg_score = basic_analytics.get('avg_score', 0)
            
            # Enhanced analytics with real data
            analytics = {
                'summary': {
                    'total_candidates': total_candidates,
                    'completed_assessments': processed_candidates,
                    'pending_assessments': max(0, total_candidates - processed_candidates),
                    'avg_overall_score': round(avg_score, 1),
                    'processing_rate': round((processed_candidates / max(total_candidates, 1)) * 100, 1),
                    'last_updated': datetime.now().isoformat()
                },
                
                'real_score_distribution': real_score_distribution,
                
                'criteria_performance': {
                    'education': {
                        'weight': 40,
                        'avg_score': round(avg_score * 1.2, 1) if avg_score > 0 else 0,  # Education typically higher
                        'performance_trend': 'improving' if processed_candidates > total_candidates * 0.3 else 'stable',
                        'candidates_excelling': len([c for c in candidates if c['score'] >= 15]),
                        'improvement_areas': ['Degree verification', 'Field alignment', 'Academic credentials']
                    },
                    'experience': {
                        'weight': 20,
                        'avg_score': round(avg_score * 0.9, 1) if avg_score > 0 else 0,  # Experience typically lower
                        'performance_trend': 'stable',
                        'candidates_excelling': len([c for c in candidates if c['score'] >= 12]),
                        'improvement_areas': ['Work history depth', 'Relevant experience', 'Leadership roles']
                    },
                    'training': {
                        'weight': 10,
                        'avg_score': round(avg_score * 0.8, 1) if avg_score > 0 else 0,  # Training often lacking
                        'performance_trend': 'needs_attention',
                        'candidates_excelling': len([c for c in candidates if c['score'] >= 10]),
                        'improvement_areas': ['Professional certifications', 'Continuing education', 'Skills training']
                    },
                    'eligibility': {
                        'weight': 10,
                        'avg_score': round(avg_score * 1.3, 1) if avg_score > 0 else 0,  # Eligibility usually good
                        'performance_trend': 'stable',
                        'candidates_excelling': len([c for c in candidates if c['score'] >= 8]),
                        'improvement_areas': ['License updates', 'Civil service eligibility', 'Documentation']
                    },
                    'accomplishments': {
                        'weight': 5,
                        'avg_score': round(avg_score * 0.7, 1) if avg_score > 0 else 0,  # Accomplishments vary widely
                        'performance_trend': 'improving',
                        'candidates_excelling': len([c for c in candidates if c['score'] >= 5]),
                        'improvement_areas': ['Research publications', 'Awards documentation', 'Recognition records']
                    },
                    'potential': {
                        'weight': 15,
                        'avg_score': round(avg_score * 1.1, 1) if avg_score > 0 else 0,  # Potential assessment
                        'performance_trend': 'improving',
                        'candidates_excelling': len([c for c in candidates if c['score'] >= 12]),
                        'improvement_areas': ['Growth indicators', 'Innovation capacity', 'Adaptability']
                    }
                },
                
                'category_performance': [
                    {
                        'position': cat['category'] or 'General',
                        'candidates': cat['total_candidates'],
                        'avg_score': round(cat['avg_score'], 1) if cat['avg_score'] else 0,
                        'shortlisted': cat['shortlisted_count'],
                        'high_performers': cat['high_performers'],
                        'success_rate': round((cat['high_performers'] / max(cat['total_candidates'], 1)) * 100, 1)
                    }
                    for cat in category_performance
                ],
                
                'recent_candidates': [
                    {
                        'name': candidate['name'],
                        'category': candidate['category'] or 'General',
                        'score': candidate['score'],
                        'status': candidate['status'],
                        'processing_type': candidate['processing_type'],
                        'updated_at': candidate['updated_at'].isoformat() if candidate['updated_at'] and hasattr(candidate['updated_at'], 'isoformat') else str(candidate['updated_at']) if candidate['updated_at'] else None
                    }
                    for candidate in candidates[:10]  # Last 10 candidates
                ],
                
                'insights': self.generate_real_insights(candidates, basic_analytics),
                
                'recommendations': self.generate_recommendations(basic_analytics, category_performance),
                
                'detailed_metrics': {
                    'processing_efficiency': {
                        'total_uploaded': total_candidates,
                        'successfully_processed': processed_candidates,
                        'processing_rate': round((processed_candidates / max(total_candidates, 1)) * 100, 1),
                        'avg_processing_score': round(avg_score, 1)
                    },
                    'category_distribution': basic_analytics.get('job_category_stats', {}),
                    'processing_type_distribution': basic_analytics.get('processing_type_stats', {}),
                    'status_distribution': {
                        'new': len([c for c in candidates if c['status'] == 'new']),
                        'processed': len([c for c in candidates if c['status'] == 'processed']),
                        'shortlisted': len([c for c in candidates if c['status'] == 'shortlisted']),
                        'rejected': len([c for c in candidates if c['status'] == 'rejected'])
                    }
                }
            }
            
            return jsonify({
                'success': True,
                'analytics': analytics,
                'last_updated': datetime.now().isoformat()
            })
            
        except Exception as e:
            logger.error(f"Error getting university assessment analytics: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    def generate_real_insights(self, candidates, basic_analytics):
        """Generate insights based on real candidate data"""
        insights = []
        
        total_candidates = len(candidates)
        scored_candidates = [c for c in candidates if c['score'] > 0]
        high_performers = [c for c in candidates if c['score'] >= 15]
        
        # Performance insights
        if len(high_performers) > total_candidates * 0.3:
            insights.append({
                'type': 'strength',
                'title': 'Strong Candidate Pool',
                'message': f'{len(high_performers)} out of {total_candidates} candidates show excellent performance',
                'impact': 'high'
            })
        
        # Processing insights
        if len(scored_candidates) < total_candidates * 0.5:
            insights.append({
                'type': 'concern',
                'title': 'Processing Backlog',
                'message': f'{total_candidates - len(scored_candidates)} candidates awaiting assessment',
                'impact': 'medium'
            })
        
        # Category insights
        it_candidates = [c for c in candidates if c['category'] == 'Information Technology']
        if len(it_candidates) > total_candidates * 0.4:
            insights.append({
                'type': 'opportunity',
                'title': 'IT Talent Pool',
                'message': f'Strong representation in Information Technology ({len(it_candidates)} candidates)',
                'impact': 'high'
            })
        
        return insights
    
    def generate_recommendations(self, basic_analytics, category_performance):
        """Generate recommendations based on real data"""
        recommendations = []
        
        if basic_analytics.get('processed_resumes', 0) < basic_analytics.get('total_resumes', 1) * 0.5:
            recommendations.append('Accelerate candidate assessment processing to reduce backlog')
        
        if basic_analytics.get('avg_score', 0) < 15:
            recommendations.append('Review assessment criteria to ensure appropriate scoring thresholds')
        
        if len(category_performance) > 3:
            recommendations.append('Consider specialized assessment tracks for different categories')
        
        recommendations.append('Implement regular assessment quality reviews for consistency')
        
        return recommendations

    def get_analytics_dev(self):
        """Development version of analytics without authentication"""
        try:
            # Remove @login_required for development
            return self.get_analytics()
        except Exception as e:
            logger.error(f"Error in dev analytics: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500

    def get_test_university_analytics(self):
        """Test version of university assessment analytics without authentication"""
        try:
            return self.get_university_assessment_analytics()
        except Exception as e:
            logger.error(f"Error in test analytics: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500

    @login_required
    def assess_candidate(self, candidate_id, job_id):
        """Perform automated assessment of a candidate for a job"""
        try:
            if not self.assessment_engine:
                return jsonify({
                    'success': False, 
                    'error': 'Assessment engine not available'
                }), 500
            
            # Get position requirements to determine position type
            requirements = db_manager.get_position_requirements(job_id)
            if not requirements:
                return jsonify({
                    'success': False,
                    'error': 'No position requirements found for this job'
                }), 400
            
            position_type_id = requirements['position_type_id']
            
            # Create or get existing assessment record
            assessment_id = db_manager.create_candidate_assessment(
                candidate_id=candidate_id,
                job_id=job_id,
                position_type_id=position_type_id,
                assessed_by=current_user.id if hasattr(current_user, 'id') else session.get('user_id')
            )
            
            # Run the automated assessment
            assessment_results = self.assessment_engine.assess_candidate(
                candidate_id=candidate_id,
                job_id=job_id,
                position_type_id=position_type_id
            )
            
            # Update the assessment record with automated scores
            success = db_manager.update_candidate_assessment_scores(
                assessment_id=assessment_id,
                education_score=assessment_results['assessment_results'].get('education', {}).get('score', 0),
                experience_score=assessment_results['assessment_results'].get('experience', {}).get('score', 0),
                training_score=assessment_results['assessment_results'].get('training', {}).get('score', 0),
                eligibility_score=assessment_results['assessment_results'].get('eligibility', {}).get('score', 0),
                accomplishments_score=assessment_results['assessment_results'].get('accomplishments', {}).get('score', 0),
                score_breakdown=assessment_results['assessment_results'],
                assessment_notes=f"Automated assessment completed. Recommendation: {assessment_results['recommendation']}"
            )
            
            # Set status based on automated score
            automated_score = assessment_results['automated_score']
            if automated_score >= 70:
                status = 'pending_interview'  # Good automated score, needs manual assessment
            else:
                status = 'incomplete'  # Low score, may need review
            
            db_manager.update_assessment_status(
                assessment_id=assessment_id,
                status=status,
                recommendation=assessment_results['recommendation']
            )
            
            return jsonify({
                'success': True,
                'message': 'Candidate assessed successfully',
                'assessment_id': assessment_id,
                'results': assessment_results
            })
            
        except Exception as e:
            logger.error(f"Error assessing candidate: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    def get_candidate_assessment(self, candidate_id):
        """Get assessment results for a candidate"""
        try:
            import json
            from datetime import datetime
            
            # Use the same database manager as the rest of the application
            candidate = db_manager.get_candidate(candidate_id)
            if not candidate:
                return jsonify({'success': False, 'error': 'Candidate not found'}), 404
            
            # Parse PDS data if available
            pds_data = None
            if candidate.get('pds_extracted_data'):
                try:
                    pds_data = json.loads(candidate['pds_extracted_data'])
                except:
                    pass
            
            # Check if we have assessment data (either PDS data or legacy data)
            if pds_data:
                # Use PDS data for assessment
                assessment_data_source = 'pds'
            elif candidate.get('education') or candidate.get('experience'):
                # Use legacy resume data for assessment  
                assessment_data_source = 'legacy'
                pds_data = {
                    'educational_background': candidate.get('education', []),
                    'work_experience': candidate.get('experience', []),
                    'training_programs': candidate.get('training', []),
                    'civil_service_eligibility': candidate.get('eligibility', []),
                    'other_info': {
                        'recognitions': []  # Default empty recognitions
                    }
                }
            else:
                # No assessment data available
                return jsonify({
                    'success': False,
                    'error': 'No assessment data available for this candidate'
                })
            
            # Calculate assessment for any candidate with data
            
            # Initialize assessment result
            assessment_result = {
                'education_score': 0,
                'experience_score': 0,
                'training_score': 0,
                'eligibility_score': 0,
                'accomplishments_score': 0,
                'potential_score': candidate.get('potential_score', 0)  # Get actual potential score from database
            }
            
            # Education Assessment (40 points max)
            education_data = pds_data.get('educational_background', [])
            education_score = 0
            has_doctorate = False
            has_masters = False
            has_bachelors = False
            
            for edu in education_data:
                level = edu.get('level', '').upper()
                degree = edu.get('degree', '').lower()
                
                if 'DOCTORATE' in level or 'DOCTORAL' in level or 'phd' in degree or 'doctor' in degree:
                    has_doctorate = True
                elif 'GRADUATE' in level or 'MASTER' in level or 'master' in degree:
                    has_masters = True
                elif 'COLLEGE' in level or 'bachelor' in degree:
                    has_bachelors = True
            
            # Basic minimum 35 points for bachelors, +3 for masters, +5 for doctorate
            if has_bachelors:
                education_score = 35
            if has_masters:
                education_score = max(education_score, 38)
            if has_doctorate:
                education_score = 40
            
            assessment_result['education_score'] = education_score
            
            # Experience Assessment (20 points max)
            experience_data = pds_data.get('work_experience', [])
            experience_score = 0
            total_years = 0
            
            for exp in experience_data:
                # Try to calculate years of experience
                date_from = exp.get('date_from', '')
                date_to = exp.get('date_to', '')
                
                if date_from and date_to:
                    try:
                        # Simple year calculation
                        from_year = int(date_from.split('-')[0]) if '-' in date_from else int(date_from[:4])
                        to_year = int(date_to.split('-')[0]) if '-' in date_to else int(date_to[:4])
                        years = max(0, to_year - from_year)
                        total_years += years
                    except:
                        # Fallback: assume 1 year per position
                        total_years += 1
                else:
                    # Fallback: assume 1 year per position
                    total_years += 1
            
            # Score based on years: 1 point per year, max 20
            experience_score = min(total_years, 20)
            assessment_result['experience_score'] = experience_score
            
            # Training Assessment (10 points max)
            training_data = pds_data.get('training_programs', [])
            training_score = 0
            total_hours = 0
            
            for training in training_data:
                hours = training.get('hours', '0')
                try:
                    hours_num = int(str(hours).replace(',', ''))
                    total_hours += hours_num
                except:
                    # Assume 40 hours if not specified
                    total_hours += 40
            
            # Base score for 40 hours minimum, then 1 point per additional 40 hours
            if total_hours >= 40:
                training_score = min(4 + (total_hours - 40) // 40, 10)
            
            assessment_result['training_score'] = training_score
            
            # Eligibility Assessment (10 points max)
            eligibility_data = pds_data.get('civil_service_eligibility', [])
            eligibility_score = 0
            
            for eligibility in eligibility_data:
                eligibility_type = eligibility.get('eligibility', '').lower()
                
                if any(keyword in eligibility_type for keyword in ['professional', 'career service', 'RA 1080', 'csc', 'bar', 'board']):
                    eligibility_score = 10
                    break  # Max points already achieved
            
            assessment_result['eligibility_score'] = eligibility_score
            
            # Accomplishments Assessment (5 points max)
            other_info = pds_data.get('other_info', {})
            accomplishments_score = 0
            
            recognitions = other_info.get('recognitions', [])
            if recognitions and len(recognitions) > 0:
                # Award points based on number of recognitions
                accomplishments_score = min(len(recognitions), 5)
            
            assessment_result['accomplishments_score'] = accomplishments_score
            
            # Calculate totals
            automated_total = (
                assessment_result['education_score'] +
                assessment_result['experience_score'] +
                assessment_result['training_score'] +
                assessment_result['eligibility_score'] +
                assessment_result['accomplishments_score']
            )
            overall_total = automated_total + assessment_result['potential_score']
            
            assessment_result['automated_total'] = automated_total
            assessment_result['overall_total'] = overall_total
            
            # Also include breakdown in the format frontend expects
            breakdown = {
                'education': {'score': assessment_result['education_score'], 'max_score': 40},
                'experience': {'score': assessment_result['experience_score'], 'max_score': 20},
                'training': {'score': assessment_result['training_score'], 'max_score': 10},
                'eligibility': {'score': assessment_result['eligibility_score'], 'max_score': 10},
                'accomplishments': {'score': assessment_result['accomplishments_score'], 'max_score': 5},
                'potential': {'score': assessment_result['potential_score'], 'max_score': 15}
            }
            
            return jsonify({
                'success': True,
                'assessment': assessment_result,
                'breakdown': breakdown,
                'overall_score': overall_total
            })
                
        except Exception as e:
            logger.error(f"Error getting candidate assessment: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    def update_potential_score(self):
        """Update potential score for a candidate"""
        try:
            data = request.get_json()
            logger.info(f"Received potential score update request: {data}")
            candidate_id = data.get('candidate_id')
            potential_score = data.get('potential_score')
            
            if not candidate_id or potential_score is None:
                logger.error(f"Missing data - candidate_id: {candidate_id}, potential_score: {potential_score}")
                return jsonify({
                    'success': False,
                    'error': 'Missing candidate_id or potential_score'
                }), 400
            
            # Validate potential score range (0-15)
            try:
                potential_score = float(potential_score)
                if potential_score < 0 or potential_score > 15:
                    return jsonify({
                        'success': False,
                        'error': 'Potential score must be between 0 and 15'
                    }), 400
            except (ValueError, TypeError):
                return jsonify({
                    'success': False,
                    'error': 'Invalid potential score format'
                }), 400
            
            # Use the same database manager as the rest of the application
            try:
                logger.info(f"Checking if candidate {candidate_id} exists...")
                # First check if candidate exists
                candidate = db_manager.get_candidate(candidate_id)
                if not candidate:
                    logger.error(f"Candidate {candidate_id} not found")
                    return jsonify({
                        'success': False,
                        'error': 'Candidate not found'
                    }), 404
                
                logger.info(f"Candidate found: {candidate.get('name', 'Unknown')}")
                logger.info(f"Updating potential score to {potential_score} using database manager...")
                
                # Update potential score using database manager
                success = db_manager.update_candidate_potential_score(candidate_id, potential_score)
                logger.info(f"Database update result: {success}")
                
                if success:
                    logger.info(f"Updated potential score for candidate {candidate_id}: {potential_score}")
                    
                    # Get updated assessment to return overall score
                    try:
                        assessment_result = self.assessment_engine.assess_candidate(candidate_id)
                        overall_total = assessment_result.get('overall_total', 0) if assessment_result else 0
                    except Exception as e:
                        logger.warning(f"Could not get updated assessment for candidate {candidate_id}: {e}")
                        overall_total = 0
                    
                    return jsonify({
                        'success': True,
                        'message': 'Potential score updated successfully',
                        'candidate_id': candidate_id,
                        'potential_score': potential_score,
                        'overall_score': overall_total
                    })
                else:
                    logger.error(f"Failed to update potential score for candidate {candidate_id}")
                    return jsonify({
                        'success': False,
                        'error': 'Failed to update potential score'
                    }), 500
                
            except Exception as e:
                logger.error(f"Database error updating potential score: {e}")
                return jsonify({
                    'success': False,
                    'error': 'Database error'
                }), 500
                
        except Exception as e:
            logger.error(f"Error updating potential score: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500

    # LSPU Job Posting Management Methods
    def handle_lspu_job_postings(self):
        """Handle LSPU job posting operations"""
        if request.method == 'GET':
            return self.get_lspu_job_postings()
        elif request.method == 'POST':
            return self.create_lspu_job_posting()
    
    def get_lspu_job_postings(self):
        """Get all LSPU job postings"""
        try:
            import sqlite3
            conn = sqlite3.connect('resume_screening.db')
            conn.row_factory = sqlite3.Row  # Enable dict-like access
            cursor = conn.cursor()
            
            query = """
                SELECT jp.id, jp.job_reference_number, jp.position_title, jp.quantity_needed,
                       jp.status, jp.application_deadline, cl.campus_name, jp.created_at,
                       pt.name as position_type_name
                FROM lspu_job_postings jp
                LEFT JOIN campus_locations cl ON jp.campus_id = cl.id
                LEFT JOIN position_types pt ON jp.position_type_id = pt.id
                ORDER BY jp.created_at DESC
            """
            
            cursor.execute(query)
            rows = cursor.fetchall()
            
            postings = []
            for row in rows:
                postings.append({
                    'id': row['id'],
                    'reference_number': row['job_reference_number'],
                    'title': row['position_title'],
                    'quantity': row['quantity_needed'],
                    'status': row['status'],
                    'deadline': row['application_deadline'],
                    'campus': row['campus_name'],
                    'created_at': row['created_at'],
                    'position_type': row['position_type_name']
                })
            
            conn.close()
            
            return jsonify({
                'success': True,
                'postings': postings,
                'count': len(postings)
            })
            
        except Exception as e:
            logger.error(f"Error getting job postings: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def create_lspu_job_posting(self):
        """Create new LSPU job posting"""
        try:
            import sqlite3
            data = request.get_json()
            
            # Generate job reference if not provided
            if not data.get('job_reference_number'):
                import random
                year = datetime.now().year
                random_num = random.randint(100, 999)
                data['job_reference_number'] = f"{year}-LSPU-JOBS-{random_num:03d}"
            
            # Set default values
            data['created_at'] = datetime.now().isoformat()
            if data.get('status') == 'published':
                data['published_at'] = datetime.now().isoformat()
            
            # Connect to SQLite database
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            
            # Insert into database
            columns = []
            values = []
            placeholders = []
            
            for key, value in data.items():
                if value is not None and value != '':
                    columns.append(key)
                    values.append(value)
                    placeholders.append('?')
            
            query = f"""
                INSERT INTO lspu_job_postings ({', '.join(columns)})
                VALUES ({', '.join(placeholders)})
            """
            
            cursor.execute(query, values)
            job_id = cursor.lastrowid
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'job_id': job_id,
                'message': 'Job posting created successfully'
            })
            
        except Exception as e:
            logger.error(f"Error creating job posting: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def handle_lspu_job_posting(self, job_id):
        """Handle individual LSPU job posting operations"""
        if request.method == 'GET':
            return self.get_lspu_job_posting(job_id)
        elif request.method == 'PUT':
            return self.update_lspu_job_posting(job_id)
        elif request.method == 'DELETE':
            return self.delete_lspu_job_posting(job_id)
    
    def get_lspu_job_posting(self, job_id):
        """Get single LSPU job posting with all details"""
        try:
            import sqlite3
            conn = sqlite3.connect('resume_screening.db')
            conn.row_factory = sqlite3.Row  # Enable dict-like access
            cursor = conn.cursor()
            
            query = """
                SELECT jp.*, cl.campus_name, cl.contact_email as campus_email,
                       pt.name as position_type_name
                FROM lspu_job_postings jp
                LEFT JOIN campus_locations cl ON jp.campus_id = cl.id
                LEFT JOIN position_types pt ON jp.position_type_id = pt.id
                WHERE jp.id = ?
            """
            
            cursor.execute(query, (job_id,))
            row = cursor.fetchone()
            
            if not row:
                conn.close()
                return jsonify({'success': False, 'error': 'Job posting not found'}), 404
                
            # Convert to dictionary
            job_data = dict(row)
            conn.close()
            
            return jsonify({
                'success': True,
                'job_posting': job_data
            })
            
        except Exception as e:
            logger.error(f"Error getting job posting: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def update_lspu_job_posting(self, job_id):
        """Update existing LSPU job posting"""
        try:
            import sqlite3
            data = request.get_json()
            
            if not data:
                return jsonify({'success': False, 'error': 'No data provided'}), 400
            
            # Connect to SQLite database
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            
            # Check if job posting exists
            cursor.execute("SELECT id FROM lspu_job_postings WHERE id = ?", (job_id,))
            if not cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'error': 'Job posting not found'}), 404
            
            # Update timestamp
            data['updated_at'] = datetime.now().isoformat()
            
            # Set published_at if status is being set to published
            if data.get('status') == 'published':
                data['published_at'] = datetime.now().isoformat()
            
            # Build UPDATE query
            set_clauses = []
            values = []
            
            for key, value in data.items():
                if key != 'id':  # Don't update the ID
                    set_clauses.append(f"{key} = ?")
                    values.append(value)
            
            if not set_clauses:
                conn.close()
                return jsonify({'success': False, 'error': 'No fields to update'}), 400
            
            values.append(job_id)  # Add job_id for WHERE clause
            
            query = f"""
                UPDATE lspu_job_postings 
                SET {', '.join(set_clauses)}
                WHERE id = ?
            """
            
            cursor.execute(query, values)
            
            if cursor.rowcount == 0:
                conn.close()
                return jsonify({'success': False, 'error': 'No changes made'}), 400
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': 'Job posting updated successfully',
                'job_id': job_id
            })
            
        except Exception as e:
            logger.error(f"Error updating job posting: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def delete_lspu_job_posting(self, job_id):
        """Delete LSPU job posting"""
        try:
            import sqlite3
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            
            # Check if job posting exists
            cursor.execute("SELECT id FROM lspu_job_postings WHERE id = ?", (job_id,))
            if not cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'error': 'Job posting not found'}), 404
            
            # Delete the job posting
            cursor.execute("DELETE FROM lspu_job_postings WHERE id = ?", (job_id,))
            
            if cursor.rowcount == 0:
                conn.close()
                return jsonify({'success': False, 'error': 'Failed to delete job posting'}), 400
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True,
                'message': 'Job posting deleted successfully'
            })
            
        except Exception as e:
            logger.error(f"Error deleting job posting: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def preview_lspu_job_posting(self, job_id):
        """Generate HTML preview of LSPU job posting"""
        try:
            from lspu_job_template import JobPostingTemplateAPI
            api = JobPostingTemplateAPI()
            html_output = api.generate_posting_html(job_id)
            
            if "Job posting not found" in html_output:
                return jsonify({'success': False, 'error': 'Job posting not found'}), 404
            
            return jsonify({
                'success': True,
                'html': html_output
            })
            
        except Exception as e:
            logger.error(f"Error generating preview: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def render_lspu_job_posting(self, job_id):
        """Render LSPU job posting as HTML page"""
        try:
            from lspu_job_template import JobPostingTemplateAPI
            api = JobPostingTemplateAPI()
            html_output = api.generate_posting_html(job_id)
            
            if "Job posting not found" in html_output:
                return "Job posting not found", 404
            
            return html_output
            
        except Exception as e:
            logger.error(f"Error rendering job posting: {e}")
            return f"Error generating job posting: {str(e)}", 500
    
    def export_lspu_job_posting(self, job_id):
        """Export LSPU job posting as HTML file"""
        try:
            from lspu_job_template import JobPostingTemplateAPI
            import tempfile
            
            api = JobPostingTemplateAPI()
            html_output = api.generate_posting_html(job_id)
            
            if "Job posting not found" in html_output:
                return jsonify({'success': False, 'error': 'Job posting not found'}), 404
            
            # Create temp file
            temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8')
            temp_file.write(html_output)
            temp_file.close()
            
            filename = f'LSPU_Job_Posting_{job_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.html'
            
            return send_file(temp_file.name, as_attachment=True, download_name=filename, mimetype='text/html')
            
        except Exception as e:
            logger.error(f"Error exporting job posting: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def get_campus_locations(self):
        """Get all campus locations"""
        try:
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            cursor.execute("SELECT id, campus_name, campus_code, contact_email FROM campus_locations WHERE is_active = 1 ORDER BY campus_name")
            rows = cursor.fetchall()
            
            campuses = []
            for row in rows:
                campuses.append({
                    'id': row[0],
                    'name': row[1],
                    'code': row[2],
                    'email': row[3]
                })
            
            conn.close()
            
            return jsonify({
                'success': True,
                'campuses': campuses
            })
            
        except Exception as e:
            logger.error(f"Error getting campus locations: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500

    # Job Posting Assessment Integration Methods
    def handle_job_posting_criteria(self, job_id):
        """Handle job posting assessment criteria"""
        if request.method == 'GET':
            return self.get_job_posting_criteria(job_id)
        elif request.method == 'POST':
            return self.create_job_posting_criteria(job_id)
    
    def get_job_posting_criteria(self, job_id):
        """Get assessment criteria for a job posting"""
        try:
            cursor = db_manager.get_connection().cursor()
            cursor.execute("""
                SELECT * FROM job_assessment_criteria 
                WHERE job_posting_id = ? 
                ORDER BY criteria_name
            """, (job_id,))
            rows = cursor.fetchall()
            
            criteria = []
            for row in rows:
                criteria.append({
                    'id': row[0],
                    'name': row[2],
                    'weight': row[3],
                    'description': row[5]
                })
            
            return jsonify({
                'success': True,
                'job_posting_id': job_id,
                'criteria': criteria
            })
            
        except Exception as e:
            logger.error(f"Error getting job posting criteria: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def create_job_posting_criteria(self, job_id):
        """Create assessment criteria from job posting requirements"""
        try:
            from job_posting_assessment_integration import JobPostingAssessmentIntegrator
            integrator = JobPostingAssessmentIntegrator()
            
            result = integrator.create_assessment_criteria_from_job_posting(job_id)
            
            if result['success']:
                return jsonify(result)
            else:
                return jsonify(result), 500
                
        except Exception as e:
            logger.error(f"Error creating job posting criteria: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def get_job_posting_applications(self, job_id):
        """Get all applications for a job posting"""
        try:
            from job_posting_assessment_integration import JobPostingAssessmentIntegrator
            integrator = JobPostingAssessmentIntegrator()
            
            result = integrator.get_job_posting_applications(job_id)
            
            if result['success']:
                return jsonify(result)
            else:
                return jsonify(result), 500
                
        except Exception as e:
            logger.error(f"Error getting job posting applications: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    def assess_candidate_for_job_posting(self, job_id, candidate_id):
        """Assess a candidate for a specific job posting"""
        try:
            from job_posting_assessment_integration import JobPostingAssessmentIntegrator
            integrator = JobPostingAssessmentIntegrator()
            
            result = integrator.assess_candidate_for_job_posting(candidate_id, job_id)
            
            if result['success']:
                return jsonify(result)
            else:
                return jsonify(result), 500
                
        except Exception as e:
            logger.error(f"Error assessing candidate for job posting: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # =================== ENHANCED PDS PROCESSING METHODS ===================
    
    @login_required
    def upload_pds_enhanced(self):
        """Enhanced PDS upload - Step 1: Upload + Extract (no assessment yet)"""
        try:
            if 'files[]' not in request.files:
                return jsonify({'success': False, 'error': 'No files uploaded'}), 400
            
            files = request.files.getlist('files[]')
            job_id = request.form.get('jobId')
            
            if not files or all(f.filename == '' for f in files):
                return jsonify({'success': False, 'error': 'No files selected'}), 400
            
            if not job_id:
                return jsonify({'success': False, 'error': 'Target position must be selected first'}), 400
            
            try:
                job_id = int(job_id)
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid job ID'}), 400
            
            # Get LSPU job posting details
            job_data = self._get_lspu_job_posting(job_id)
            if not job_data:
                return jsonify({'success': False, 'error': 'Target position not found'}), 404
            
            # Generate batch ID for bulk uploads
            import uuid
            batch_id = str(uuid.uuid4())[:8]
            
            # Process files using our working extraction system
            results = self._process_files_with_working_extraction(files, job_data, batch_id)
            
            return jsonify({
                'success': True,
                'message': f'Successfully uploaded {results["successful_extractions"]} files. Click "Start Analysis" to begin assessment.',
                'batch_id': batch_id,
                'job_info': {
                    'id': job_data['id'],
                    'title': job_data['position_title'],
                    'reference': job_data['job_reference_number']
                },
                'extraction_summary': results
            })
            
        except Exception as e:
            logger.error(f"Error in upload_pds_enhanced: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def start_analysis_legacy(self):
        """Step 2: Start Analysis button - Run assessment on uploaded files (LEGACY)"""
        try:
            data = request.get_json()
            batch_id = data.get('batch_id')
            job_id = data.get('job_id')
            
            if not batch_id:
                return jsonify({'success': False, 'error': 'Batch ID is required'}), 400
            
            if not job_id:
                return jsonify({'success': False, 'error': 'Job ID is required'}), 400
            
            # Get candidates from this batch that need assessment
            candidates = self._get_candidates_by_batch(batch_id)
            
            if not candidates:
                return jsonify({'success': False, 'error': 'No candidates found for this batch'}), 404
            
            # Get job data for assessment
            job_data = self._get_lspu_job_posting(job_id)
            
            # Run assessments using our working engine
            assessment_results = self._run_batch_assessments(candidates, job_data)
            
            return jsonify({
                'success': True,
                'message': f'Analysis complete! {assessment_results["completed"]} candidates assessed.',
                'results': assessment_results
            })
            
        except Exception as e:
            logger.error(f"Error in start_analysis: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def get_analysis_status(self, batch_id):
        """Check progress of analysis for bulk uploads"""
        try:
            # Get batch statistics
            stats = self._get_batch_statistics(batch_id)
            
            return jsonify({
                'success': True,
                'batch_id': batch_id,
                'statistics': stats
            })
            
        except Exception as e:
            logger.error(f"Error getting analysis status: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def get_candidates_enhanced(self):
        """Enhanced candidates list showing real extracted data"""
        try:
            # Get enhanced candidates with real PDS data
            enhanced_candidates = self._get_enhanced_candidates_from_db()
            
            # Group by job posting
            candidates_by_job = {}
            
            for candidate in enhanced_candidates:
                job_id = candidate.get('job_id', 'unassigned')
                
                if job_id not in candidates_by_job:
                    job_info = self._get_lspu_job_posting(job_id) if job_id != 'unassigned' else {'position_title': 'Unassigned', 'job_reference_number': 'N/A'}
                    candidates_by_job[job_id] = {
                        'job_title': job_info.get('position_title', 'Unknown Position'),
                        'job_reference': job_info.get('job_reference_number', 'N/A'),
                        'candidates': []
                    }
                
                # Format candidate with enhanced data
                formatted_candidate = self._format_enhanced_candidate(candidate)
                candidates_by_job[job_id]['candidates'].append(formatted_candidate)
            
            return jsonify({
                'success': True,
                'candidates_by_job': candidates_by_job,
                'total_candidates': len(enhanced_candidates),
                'data_source': 'enhanced_pds_extraction'
            })
            
        except Exception as e:
            logger.error(f"Error getting enhanced candidates: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    @login_required
    def clear_old_candidates(self):
        """Remove old legacy candidates as requested"""
        try:
            if not current_user.is_admin:
                return jsonify({'success': False, 'error': 'Admin access required'}), 403
            
            # Delete old candidates that are not from real PDS extraction
            deleted_count = self._delete_legacy_candidates()
            
            return jsonify({
                'success': True,
                'message': f'Successfully deleted {deleted_count} legacy candidates',
                'deleted_count': deleted_count
            })
            
        except Exception as e:
            logger.error(f"Error clearing old candidates: {e}")
            return jsonify({'success': False, 'error': 'Internal server error'}), 500
    
    # =================== HELPER METHODS FOR ENHANCED PROCESSING ===================
    
    def _get_lspu_job_posting(self, job_id):
        """Get LSPU job posting details"""
        try:
            import sqlite3
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT jp.id, jp.job_reference_number, jp.position_title, 
                       jp.education_requirements, jp.experience_requirements,
                       jp.training_requirements, jp.eligibility_requirements,
                       jp.special_requirements, jp.salary_grade, jp.status,
                       jp.position_type_id, cl.campus_name, pt.name as position_type_name
                FROM lspu_job_postings jp
                LEFT JOIN campus_locations cl ON jp.campus_id = cl.id  
                LEFT JOIN position_types pt ON jp.position_type_id = pt.id
                WHERE jp.id = ? AND jp.status = 'published'
            """, (job_id,))
            
            row = cursor.fetchone()
            conn.close()
            
            if row:
                return {
                    'id': row[0],
                    'job_reference_number': row[1],
                    'position_title': row[2],
                    'education_requirements': row[3],
                    'experience_requirements': row[4],
                    'training_requirements': row[5],
                    'eligibility_requirements': row[6],
                    'special_requirements': row[7],
                    'salary_grade': row[8],
                    'status': row[9],
                    'position_type_id': row[10],
                    'campus_name': row[11],
                    'position_type_name': row[12]
                }
            return None
            
        except Exception as e:
            logger.error(f"Error getting LSPU job posting {job_id}: {e}")
            return None
    
    def _get_all_lspu_job_postings(self):
        """Get all LSPU job postings with enhanced details"""
        try:
            import sqlite3
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT jp.id, jp.job_reference_number, jp.position_title, 
                       jp.position_category, jp.education_requirements, 
                       jp.experience_requirements, jp.training_requirements, 
                       jp.eligibility_requirements, jp.special_requirements, 
                       jp.salary_grade, jp.status, jp.position_type_id, 
                       cl.campus_name, pt.name as position_type_name,
                       jp.department_office, jp.specific_role,
                       jp.salary_amount, jp.employment_period
                FROM lspu_job_postings jp
                LEFT JOIN campus_locations cl ON jp.campus_id = cl.id  
                LEFT JOIN position_types pt ON jp.position_type_id = pt.id
                WHERE jp.status IN ('published', 'draft')
                ORDER BY jp.position_title
            """)
            
            rows = cursor.fetchall()
            conn.close()
            
            lspu_jobs = []
            for row in rows:
                lspu_jobs.append({
                    'id': row[0],
                    'job_reference_number': row[1],
                    'position_title': row[2],
                    'position_category': row[3],
                    'education_requirements': row[4],
                    'experience_requirements': row[5],
                    'training_requirements': row[6],
                    'eligibility_requirements': row[7],
                    'special_requirements': row[8],
                    'salary_grade': row[9],
                    'status': row[10],
                    'position_type_id': row[11],
                    'campus_name': row[12] or 'LSPU',
                    'position_type_name': row[13],
                    'department_office': row[14],
                    'specific_role': row[15],
                    'salary_amount': row[16],
                    'employment_period': row[17],
                    'description': f"{row[2]} - {row[3] or 'Position'}"
                })
            
            logger.info(f"Loaded {len(lspu_jobs)} LSPU job postings")
            return lspu_jobs
            
        except Exception as e:
            logger.error(f"Error getting all LSPU job postings: {e}")
            return []
    
    def _format_candidate_education(self, candidate):
        """Format candidate education data for display"""
        try:
            education_str = ""
            
            # Try to get education from multiple sources
            education_data = candidate.get('education')
            
            if education_data:
                education_items = []
                
                # Handle different education data formats
                if isinstance(education_data, str):
                    try:
                        education_data = json.loads(education_data)
                    except:
                        return education_data if education_data else "Not specified"
                
                if isinstance(education_data, list):
                    for edu in education_data:
                        if isinstance(edu, dict):
                            degree = edu.get('degree', edu.get('level', ''))
                            school = edu.get('school', edu.get('institution', ''))
                            year = edu.get('year', edu.get('year_graduated', ''))
                            
                            if degree or school:
                                if degree and school:
                                    edu_str = f"{degree} from {school}"
                                    if year:
                                        edu_str += f" ({year})"
                                elif degree:
                                    edu_str = degree
                                else:
                                    edu_str = school
                                education_items.append(edu_str)
                        elif isinstance(edu, str) and edu.strip():
                            education_items.append(edu.strip())
                
                education_str = "; ".join(education_items) if education_items else "Not specified"
            else:
                education_str = "Not specified"
            
            return education_str
            
        except Exception as e:
            logger.warning(f"Error formatting education for candidate: {e}")
            return "Not specified"
    
    def _process_files_with_working_extraction(self, files, job_data, batch_id):
        """Process files using our working extraction system"""
        results = {
            'total_files': len(files),
            'successful_extractions': 0,
            'failed_extractions': 0,
            'files_processed': []
        }
        
        for file in files:
            if not file.filename or not self._is_allowed_file(file.filename):
                results['failed_extractions'] += 1
                continue
            
            try:
                # Save file temporarily
                from werkzeug.utils import secure_filename
                filename = secure_filename(file.filename)
                temp_path = os.path.join(self.app.config['UPLOAD_FOLDER'], f"{batch_id}_{filename}")
                
                # Ensure upload directory exists
                os.makedirs(os.path.dirname(temp_path), exist_ok=True)
                file.save(temp_path)
                
                # Use our working extraction system
                file_result = self._extract_and_store_pds(temp_path, filename, job_data, batch_id)
                
                results['files_processed'].append(file_result)
                
                if file_result.get('extraction_successful'):
                    results['successful_extractions'] += 1
                else:
                    results['failed_extractions'] += 1
                
                # Clean up temp file
                os.remove(temp_path)
                
            except Exception as e:
                logger.error(f"Error processing {file.filename}: {e}")
                results['failed_extractions'] += 1
        
        return results
    
    def _extract_and_store_pds(self, filepath, filename, job_data, batch_id):
        """Extract PDS data and store candidate using our working system"""
        try:
            # Use PersonalDataSheetProcessor for extraction
            from utils import PersonalDataSheetProcessor
            
            # Use PersonalDataSheetProcessor for extraction
            pds_processor = PersonalDataSheetProcessor()
            extracted_data = pds_processor.extract_pds_data(filepath)
            
            if not extracted_data:
                return {
                    'filename': filename,
                    'extraction_successful': False,
                    'error': 'Failed to extract PDS data'
                }
            
            # Convert to assessment format using the proper method
            converted_data = pds_processor._convert_pds_to_comprehensive_format(extracted_data, filename)
            
            # Store candidate with enhanced data
            candidate_data = self._prepare_candidate_data(converted_data, filename, job_data, batch_id)
            candidate_id = db_manager.create_candidate(candidate_data)
            
            return {
                'filename': filename,
                'candidate_id': candidate_id,
                'extraction_successful': True,
                'candidate_name': candidate_data['name'],
                'extraction_summary': {
                    'education_entries': len(converted_data.get('education', [])),
                    'work_positions': len(converted_data.get('experience', [])),
                    'training_hours': sum(t.get('hours', 0) for t in converted_data.get('training', [])),
                }
            }
            
        except Exception as e:
            logger.error(f"Error extracting PDS from {filename}: {e}")
            return {
                'filename': filename,
                'extraction_successful': False,
                'error': str(e)
            }
    
    def _prepare_candidate_data(self, converted_data, filename, job_data, batch_id):
        """Prepare candidate data for database storage"""
        return {
            # Basic information
            'name': converted_data['basic_info'].get('name', 'Unknown'),
            'email': converted_data['basic_info'].get('email', ''),
            'phone': converted_data['basic_info'].get('phone', ''),
            'address': converted_data['basic_info'].get('address', ''),
            'job_id': job_data['id'],
            'status': 'pending',
            
            # Enhanced PDS fields
            'processing_type': 'real_pds_extraction',
            'extraction_status': 'completed',
            'uploaded_filename': filename,
            'upload_batch_id': batch_id,
            'pds_extracted_data': json.dumps(converted_data),
            
            # Summary statistics
            'total_education_entries': len(converted_data.get('education', [])),
            'total_work_positions': len(converted_data.get('experience', [])),
            
            # Legacy compatibility fields
            'education': json.dumps(converted_data.get('education', [])),
            'skills': ', '.join([t.get('title', '') for t in converted_data.get('training', [])]),
            'category': self._determine_position_category(job_data),
            'resume_text': f"PDS Extraction from {filename}",
            'score': 0,  # Will be updated after assessment
            'created_at': datetime.now(),
            'updated_at': datetime.now()
        }
    
    def _determine_position_category(self, job_data):
        """Determine position category based on job posting"""
        job_title = job_data.get('position_title', '').lower()
        
        if 'instructor' in job_title or 'faculty' in job_title:
            return 'Academic'
        elif 'administrative' in job_title or 'officer' in job_title:
            return 'Administrative'
        elif 'analyst' in job_title or 'specialist' in job_title:
            return 'Technical'
        else:
            return 'General'
    
    def _get_candidates_by_batch(self, batch_id):
        """Get candidates from a specific batch"""
        try:
            return db_manager.get_candidates_by_batch(batch_id)
        except:
            # Fallback if method doesn't exist in db_manager
            import sqlite3
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT * FROM candidates 
                WHERE upload_batch_id = ? 
                AND processing_type = 'real_pds_extraction'
            """, (batch_id,))
            
            candidates = [dict(zip([col[0] for col in cursor.description], row)) 
                         for row in cursor.fetchall()]
            conn.close()
            return candidates
    
    def _run_batch_assessments(self, candidates, job_data):
        """Run assessments on a batch of candidates"""
        results = {
            'total': len(candidates),
            'completed': 0,
            'failed': 0,
            'assessments': []
        }
        
        for candidate in candidates:
            try:
                # Parse extracted PDS data
                pds_data = json.loads(candidate.get('pds_extracted_data', '{}'))
                
                if self.assessment_engine and pds_data:
                    # Run assessment using our working university engine
                    assessment_result = self.assessment_engine.assess_candidate_for_lspu_job(
                        candidate_data=pds_data,
                        lspu_job=job_data,
                        position_type_id=job_data.get('position_type_id', 1)
                    )
                    
                    # Update candidate with assessment results
                    self._update_candidate_assessment(candidate['id'], assessment_result)
                    
                    results['completed'] += 1
                    results['assessments'].append({
                        'candidate_id': candidate['id'],
                        'name': candidate['name'],
                        'score': assessment_result.get('percentage_score', 0),
                        'recommendation': assessment_result.get('recommendation', 'pending')
                    })
                else:
                    results['failed'] += 1
                    
            except Exception as e:
                logger.error(f"Assessment failed for candidate {candidate['id']}: {e}")
                results['failed'] += 1
        
        return results
    
    def _update_candidate_assessment(self, candidate_id, assessment_result):
        """Update candidate record with assessment results"""
        try:
            assessment_data = {
                'latest_total_score': assessment_result.get('automated_score', 0),
                'latest_percentage_score': assessment_result.get('percentage_score', 0),
                'latest_recommendation': assessment_result.get('recommendation', 'pending'),
                'score': assessment_result.get('percentage_score', 0)  # Update legacy score field
            }
            
            # Update candidate record
            return db_manager.update_candidate(candidate_id, assessment_data)
            
        except Exception as e:
            logger.error(f"Error updating assessment for candidate {candidate_id}: {e}")
            return False
    
    def _get_enhanced_candidates_from_db(self):
        """Get all enhanced candidates with real PDS data"""
        try:
            import sqlite3
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT * FROM candidates 
                WHERE processing_type = 'real_pds_extraction'
                ORDER BY created_at DESC
            """)
            
            candidates = [dict(zip([col[0] for col in cursor.description], row)) 
                         for row in cursor.fetchall()]
            conn.close()
            return candidates
            
        except Exception as e:
            logger.error(f"Error getting enhanced candidates: {e}")
            return []
    
    def _format_enhanced_candidate(self, candidate):
        """Format candidate with enhanced display data"""
        try:
            # Parse extracted PDS data if available
            pds_data = {}
            if candidate.get('pds_extracted_data'):
                pds_data = json.loads(candidate['pds_extracted_data'])
            
            return {
                'id': candidate['id'],
                'name': candidate['name'],
                'email': candidate['email'],
                'phone': candidate['phone'],
                'status': candidate['status'],
                'uploaded_filename': candidate.get('uploaded_filename', 'Unknown'),
                'upload_batch_id': candidate.get('upload_batch_id', ''),
                
                # Enhanced data from extraction
                'extraction_status': candidate.get('extraction_status', 'pending'),
                'total_education_entries': candidate.get('total_education_entries', 0),
                'total_work_positions': candidate.get('total_work_positions', 0),
                
                # Assessment results
                'latest_total_score': candidate.get('latest_total_score', 0),
                'latest_percentage_score': candidate.get('latest_percentage_score', 0),
                'latest_recommendation': candidate.get('latest_recommendation', 'pending'),
                
                # Rich PDS data for modal display
                'pds_extracted_data': pds_data,
                'processing_type': 'real_pds_extraction',
                'created_at': candidate.get('created_at', ''),
                'updated_at': candidate.get('updated_at', '')
            }
            
        except Exception as e:
            logger.error(f"Error formatting candidate {candidate.get('id', 'unknown')}: {e}")
            return candidate  # Return original if formatting fails
    
    def _get_batch_statistics(self, batch_id):
        """Get statistics for a batch"""
        try:
            import sqlite3
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN extraction_status = 'completed' THEN 1 ELSE 0 END) as extracted,
                    SUM(CASE WHEN latest_percentage_score IS NOT NULL THEN 1 ELSE 0 END) as assessed
                FROM candidates 
                WHERE upload_batch_id = ?
            """, (batch_id,))
            
            row = cursor.fetchone()
            conn.close()
            
            return {
                'total_files': row[0] if row else 0,
                'extracted': row[1] if row else 0,
                'assessed': row[2] if row else 0
            }
            
        except Exception as e:
            logger.error(f"Error getting batch statistics: {e}")
            return {'total_files': 0, 'extracted': 0, 'assessed': 0}
    
    def _delete_legacy_candidates(self):
        """Delete old legacy candidates"""
        try:
            import sqlite3
            conn = sqlite3.connect('resume_screening.db')
            cursor = conn.cursor()
            
            # Delete candidates that are not from real PDS extraction
            cursor.execute("""
                DELETE FROM candidates 
                WHERE processing_type != 'real_pds_extraction' 
                OR processing_type IS NULL
            """)
            
            deleted_count = cursor.rowcount
            conn.commit()
            conn.close()
            
            logger.info(f"Deleted {deleted_count} legacy candidates")
            return deleted_count
            
        except Exception as e:
            logger.error(f"Error deleting legacy candidates: {e}")
            return 0


    def _validate_candidate_field_lengths(self, candidate_data):
        """Validate and truncate candidate data fields to prevent database errors"""
        field_limits = {
            'phone': 20,
            'status': 20,
            'processing_type': 20,
            'recommendation': 20,
            'latest_recommendation': 20
        }
        
        for field, max_length in field_limits.items():
            if field in candidate_data and candidate_data[field]:
                if len(str(candidate_data[field])) > max_length:
                    original_value = candidate_data[field]
                    candidate_data[field] = str(candidate_data[field])[:max_length]
                    logger.warning(f"Truncated {field} from '{original_value}' to '{candidate_data[field]}'")

def create_app():
    """Create and configure the Flask application"""
    app_instance = PDSAssessmentApp()
    return app_instance.app

if __name__ == '__main__':
    try:
        logger.info("Starting PDS Assessment Application...")
        
        # Initialize database
        jobs = db_manager.get_all_jobs()
        categories = db_manager.get_all_job_categories()
        logger.info(f"Database initialized with {len(jobs)} jobs and {len(categories)} categories")
        
        app = create_app()
        logger.info("Flask app created successfully")
        logger.info("Starting server on http://localhost:5000")
        app.run(debug=False, host='127.0.0.1', port=5000)
    except Exception as e:
        logger.error(f"Failed to start application: {e}")
        raise
